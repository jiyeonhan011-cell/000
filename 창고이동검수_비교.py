"""
창고이동검수 비교 자동화 프로그램 v2
- 기준 파일 (XLS): I열(비고)에서 'TCT 케이터링 이동처리' 필터 → F열(이동수량) 수집
- 비교 파일 (XLSX): 작업내역 K열(수량) ÷ 2
- 품목명 기준 매칭 후 일치/불일치 판별
- 납품처(입고창고) 단위로 상세 결과 출력

사용법:
  python3 창고이동검수_비교.py <창고이동.xls> <작업내역.xlsx>
"""

import sys, os, re
import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── 입고창고 ↔ 스케줄명 매핑 ─────────────────────────────────────────────────
# XLS 입고창고명(key) → {"sched": [스케줄명 패턴], "erp": "ERP명 포함 문자열(선택)"}
# erp 지정 시: 해당 스케줄 중 ERP명(U열)이 일치하는 행만 필터
WAREHOUSE_SCHEDULE_MAP = {
    "에스피씨지에프에스":    {"sched": ["SPC_"]},
    "동원홈푸드":            {"sched": ["동원_"]},
    "현대그린푸드":          {"sched": ["현대_"]},
    "딜리버리랩":            {"sched": ["딜리버리랩"]},
    "신세계푸드":            {"sched": ["신세계"]},
    "씨제이프레시웨이":      {"sched": ["CJ_"]},
    # 아워홈: 스케줄명은 같지만 ERP명(영업소)으로 각각 분리
    "아워홈 동서울영업소":   {"sched": ["아워홈1_"], "erp": "동서울영업소"},
    "아워홈 안산영업소":     {"sched": ["아워홈1_"], "erp": "안산영업소"},
    "아워홈 용인2영업소":    {"sched": ["아워홈1_"], "erp": "용인2영업소"},
    "삼성웰스토리 평택":     {"sched": ["웰스토리1_평택"]},
    "삼성웰스토리 용인":     {"sched": ["웰스토리1_용인"]},
    "삼성웰스토리 오산":     {"sched": ["웰스토리1_오산"]},
    "삼성웰스토리 왜관":     {"sched": ["웰스토리1_왜관"]},
    "삼성웰스토리 광주":     {"sched": ["웰스토리1_광주"]},
    "삼성웰스토리 김해":     {"sched": ["웰스토리1_김해"]},
    "푸드머스":              {"sched": ["푸드머스1"]},
    "푸디스트":              {"sched": ["한화_푸디스트_"]},
    "비젼유통":              {"sched": ["비전유통"]},
    # 스마트푸드 3센터 → 아모제
    "스마트푸드 오산센터":   {"sched": ["아모제"]},
    "스마트푸드 영남센터":   {"sched": ["아모제"]},
    "스마트푸드 호남센터":   {"sched": ["아모제"]},
    "캘리스코":              {"sched": []},  # XLSX에 없음
}

# ── 스타일 ────────────────────────────────────────────────────────────────────
def _side(): return Side(style="thin")
def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())

H_FILL   = PatternFill("solid", fgColor="1F4E79")   # 진파랑
T_FILL   = PatternFill("solid", fgColor="2E75B6")   # 파랑
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")   # 연파랑
OK_FILL  = PatternFill("solid", fgColor="C6EFCE")   # 녹색
ERR_FILL = PatternFill("solid", fgColor="FFC7CE")   # 빨강
WARN_FILL= PatternFill("solid", fgColor="FFEB9C")   # 노랑
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")   # 연회색

W_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
B_FONT = Font(name="맑은 고딕", bold=True, size=10)
N_FONT = Font(name="맑은 고딕", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")


def cell(ws, r, c, v, font=None, fill=None, align=None, fmt=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font   = font  or N_FONT
    cl.fill   = fill  or PatternFill()
    cl.alignment = align or LEFT
    cl.border = _border()
    if fmt: cl.number_format = fmt
    return cl


def col_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── 데이터 로드 ───────────────────────────────────────────────────────────────

def load_xls(path: str) -> dict:
    """
    XLS → {입고창고: {품목명: {'수량': float, '품목코드': str, '날짜': str}}}
    I열(비고)에서 'TCT 케이터링 이동처리' 필터, F열(이동수량) 사용
    """
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    data: dict[str, dict] = defaultdict(lambda: defaultdict(lambda: {"수량": 0.0, "품목코드": "", "날짜": ""}))
    count = 0
    for i in range(1, ws.nrows):
        note = str(ws.cell_value(i, 8))          # I열 = index 8
        if "케이터링 이동처리" not in note:
            continue
        warehouse = str(ws.cell_value(i, 21)).strip()   # 입고창고
        item_code = str(ws.cell_value(i, 9)).strip()    # 품목코드
        item_name = str(ws.cell_value(i, 10)).strip()   # 품목명
        qty       = float(ws.cell_value(i, 5) or 0)    # F열 이동수량
        date      = str(ws.cell_value(i, 3))            # 이동일자
        d = data[warehouse][item_name]
        d["수량"]    += qty
        d["품목코드"] = item_code
        d["날짜"]     = date
        count += 1
    print(f"  XLS: TCT 케이터링 이동처리 {count}행, 입고창고 {len(data)}개")
    return dict(data)


def load_xlsx(path: str) -> list:
    """
    XLSX → 행 리스트: [{"sched", "erp", "name", "qty"}]
    K열(수량), F열(스케줄명), U열(ERP명) 수집
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        sched = str(row[5]).strip()  if row[5]  else ""
        name  = str(row[7]).strip()  if row[7]  else ""
        erp   = str(row[20]).strip() if row[20] else ""
        qty   = float(row[10] or 0)
        if sched and name:
            rows.append({"sched": sched, "erp": erp, "name": name, "qty": qty})
    print(f"  XLSX: {len(rows)}행 로드")
    return rows


def build_xlsx_by_warehouse(xlsx_rows: list) -> dict:
    """XLSX 행 목록을 XLS 입고창고 기준으로 집계
    erp 필터가 있으면 ERP명 포함 여부로 추가 필터링"""
    result: dict[str, dict] = {}
    for wh, cfg in WAREHOUSE_SCHEDULE_MAP.items():
        patterns   = cfg["sched"]
        erp_filter = cfg.get("erp", "")
        merged: dict[str, float] = defaultdict(float)
        matched_scheds = set()
        for r in xlsx_rows:
            sched_match = any(r["sched"].startswith(p) or p in r["sched"] for p in patterns) if patterns else False
            if not sched_match:
                continue
            if erp_filter and erp_filter not in r["erp"]:
                continue
            merged[r["name"]] += r["qty"]
            matched_scheds.add(r["sched"] + (f"[{erp_filter}]" if erp_filter else ""))
        result[wh] = {
            "items":  dict(merged),
            "scheds": sorted(matched_scheds),
        }
    return result


# ── 비교 로직 ─────────────────────────────────────────────────────────────────

def fmt_qty(v):
    return int(v) if isinstance(v, float) and v == int(v) else round(v, 2)


def compare(xls_data: dict, xlsx_by_wh: dict) -> dict:
    """
    납품처별 품목 단위 비교
    반환: {입고창고: {'match': [...], 'mismatch': [...], 'xls_only': [...], 'xlsx_only': [...]}}
    """
    results = {}
    all_warehouses = set(xls_data.keys()) | set(xlsx_by_wh.keys())

    for wh in sorted(all_warehouses):
        xls_items  = xls_data.get(wh, {})
        xlsx_info  = xlsx_by_wh.get(wh, {"items": {}, "scheds": []})
        xlsx_items = xlsx_info["items"]
        scheds     = xlsx_info["scheds"]

        match, mismatch, xls_only, xlsx_only = [], [], [], []

        all_names = set(xls_items.keys()) | set(xlsx_items.keys())
        for name in sorted(all_names):
            xls_entry = xls_items.get(name, None)
            xlsx_qty  = xlsx_items.get(name, None)

            # xls_entry는 {"수량": float, "품목코드": str, "날짜": str} 형태
            xls_qty  = xls_entry["수량"] if xls_entry is not None else None
            item_code = xls_entry["품목코드"] if xls_entry is not None else ""

            if xls_qty is not None and xlsx_qty is not None:
                xlsx_half = xlsx_qty / 2
                xls_f = fmt_qty(xls_qty)
                half_f = fmt_qty(xlsx_half)
                orig_f = fmt_qty(xlsx_qty)
                diff   = round(xls_qty - xlsx_half, 4)
                item = {
                    "품목명": name,
                    "품목코드": item_code,
                    "XLS이동수량": xls_f,
                    "XLSX원본수량": orig_f,
                    "XLSX÷2": half_f,
                    "차이": fmt_qty(diff),
                }
                if diff == 0:
                    match.append(item)
                else:
                    reason = ""
                    if diff > 0:
                        reason = f"XLS 수량이 {abs(fmt_qty(diff))} 더 많음"
                    else:
                        reason = f"XLSX÷2 수량이 {abs(fmt_qty(diff))} 더 많음"
                    item["불일치사유"] = reason
                    mismatch.append(item)
            elif xls_qty is not None:
                xls_only.append({"품목명": name, "XLS이동수량": fmt_qty(xls_qty), "불일치사유": "작업내역 파일에 없음"})
            else:
                xlsx_half = xlsx_qty / 2
                xlsx_only.append({"품목명": name, "XLSX원본수량": fmt_qty(xlsx_qty),
                                   "XLSX÷2": fmt_qty(xlsx_half), "불일치사유": "창고이동 파일에 없음"})

        results[wh] = {
            "match":     match,
            "mismatch":  mismatch,
            "xls_only":  xls_only,
            "xlsx_only": xlsx_only,
            "scheds":    scheds,
        }
    return results


# ── Excel 출력 ────────────────────────────────────────────────────────────────

def write_excel(results: dict, xls_data: dict, output_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("전체요약")
    _write_summary(ws_sum, results)

    for wh, res in results.items():
        sheet_name = wh[:31]
        ws = wb.create_sheet(sheet_name)
        _write_detail(ws, wh, res)

    wb.save(output_path)
    print(f"\n✅ 결과 저장: {output_path}")


def _write_summary(ws, results: dict):
    ws.merge_cells("A1:I1")
    c = ws.cell(1, 1, "창고이동 검수 비교 결과 요약  (기준: XLS 이동수량  vs  작업내역 수량 ÷ 2)")
    c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
    c.fill = H_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    hdrs = ["납품처(입고창고)", "매핑 스케줄명", "일치", "불일치", "XLS만", "XLSX만", "전체품목", "일치율", "판정"]
    for col, h in enumerate(hdrs, 1):
        cell(ws, 2, col, h, font=W_FONT, fill=T_FILL, align=CENTER)
    ws.row_dimensions[2].height = 22

    row = 3
    for wh, res in results.items():
        n_match = len(res["match"])
        n_mm    = len(res["mismatch"])
        n_xo    = len(res["xls_only"])
        n_xlsx  = len(res["xlsx_only"])
        total   = n_match + n_mm + n_xo + n_xlsx
        rate    = round(n_match / total * 100, 1) if total else 0
        verdict = "✔ 전체일치" if (n_mm == 0 and n_xo == 0 and n_xlsx == 0) else ("⚠ 부분일치" if n_match > 0 else "✘ 불일치")
        fill    = OK_FILL if verdict.startswith("✔") else (WARN_FILL if verdict.startswith("⚠") else ERR_FILL)

        scheds_str = ", ".join(res["scheds"]) if res["scheds"] else "(매핑없음)"
        vals = [wh, scheds_str, n_match, n_mm, n_xo, n_xlsx, total, f"{rate}%", verdict]
        for col, v in enumerate(vals, 1):
            cell(ws, row, col, v, fill=fill, align=CENTER if col != 2 else LEFT)
        row += 1

    col_width(ws, [22, 45, 8, 8, 8, 8, 8, 8, 12])


def _write_detail(ws, wh: str, res: dict):
    ws.merge_cells("A1:G1")
    scheds_str = ", ".join(res["scheds"]) if res["scheds"] else "(XLSX 매핑 없음)"
    title = f"[{wh}]  ←→  스케줄: {scheds_str}"
    c = ws.cell(1, 1, title)
    c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    c.fill = H_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # 통계 행
    n_m = len(res["match"]); n_mm = len(res["mismatch"])
    n_xo = len(res["xls_only"]); n_xlsx = len(res["xlsx_only"])
    total = n_m + n_mm + n_xo + n_xlsx
    rate = round(n_m / total * 100, 1) if total else 0
    ws.merge_cells("A2:G2")
    stat = f"일치: {n_m}  |  불일치: {n_mm}  |  XLS만: {n_xo}  |  XLSX만: {n_xlsx}  |  전체: {total}  |  일치율: {rate}%"
    c2 = ws.cell(2, 1, stat)
    c2.font = B_FONT; c2.fill = SUB_FILL; c2.alignment = LEFT
    ws.row_dimensions[2].height = 18

    # 헤더
    hdrs = ["구분", "품목명", "XLS 이동수량\n(창고이동)", "XLSX 원본수량\n(작업내역)", "XLSX÷2\n(비교기준)", "차이\n(XLS-XLSX÷2)", "불일치 사유"]
    for col, h in enumerate(hdrs, 1):
        cell(ws, 3, col, h, font=W_FONT, fill=T_FILL, align=CENTER)
    ws.row_dimensions[3].height = 30

    r = 4

    def write_section(items, label, fill, extra_fill=None):
        nonlocal r
        if not items: return
        ws.merge_cells(f"A{r}:G{r}")
        sec = ws.cell(r, 1, f"▶ {label} ({len(items)}건)")
        sec.font = B_FONT; sec.fill = T_FILL; sec.alignment = LEFT
        ws.row_dimensions[r].height = 16
        r += 1
        for idx, it in enumerate(items):
            base = extra_fill if (extra_fill and idx % 2 == 0) else fill
            xls_q  = it.get("XLS이동수량", "")
            xlsx_q = it.get("XLSX원본수량", "")
            half_q = it.get("XLSX÷2", "")
            diff   = it.get("차이", "")
            reason = it.get("불일치사유", "")
            vals = [label.split()[0], it["품목명"], xls_q, xlsx_q, half_q, diff, reason]
            for col, v in enumerate(vals, 1):
                f = fill if col in (6, 7) and reason else base
                cell(ws, r, col, v, fill=f, align=CENTER if col in (1,3,4,5,6) else LEFT)
            r += 1

    write_section(res["match"],    "✔ 일치",        OK_FILL,   ALT_FILL)
    write_section(res["mismatch"], "✘ 불일치 (수량차이)", ERR_FILL)
    write_section(res["xls_only"],"△ XLS에만 있음 (작업내역 누락)", WARN_FILL)
    write_section(res["xlsx_only"],"▽ XLSX에만 있음 (창고이동 누락)", WARN_FILL)

    col_width(ws, [10, 40, 14, 14, 12, 12, 30])


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        # 같은 폴더에서 자동 탐색
        files = os.listdir(".")
        xls_files  = [f for f in files if f.endswith(".xls") and not f.endswith(".xlsx")]
        xlsx_files = [f for f in files if f.endswith(".xlsx") and "검수결과" not in f]
        if not xls_files or not xlsx_files:
            print("사용법: python3 창고이동검수_비교.py <창고이동.xls> <작업내역.xlsx>")
            sys.exit(1)
        xls_path, xlsx_path = xls_files[0], xlsx_files[0]
        print(f"파일 자동 감지: {xls_path}, {xlsx_path}")
    else:
        xls_path  = sys.argv[1]
        xlsx_path = sys.argv[2]

    for p in [xls_path, xlsx_path]:
        if not os.path.exists(p):
            print(f"❌ 파일 없음: {p}"); sys.exit(1)

    base = os.path.splitext(os.path.basename(xls_path))[0]
    out  = os.path.join(os.path.dirname(xls_path) or ".", f"{base}_검수비교결과.xlsx")

    print(f"\n📂 XLS  로딩: {xls_path}")
    xls_data = load_xls(xls_path)

    print(f"📂 XLSX 로딩: {xlsx_path}")
    xlsx_data = load_xlsx(xlsx_path)

    print("\n🔗 납품처-스케줄 매핑 중...")
    xlsx_by_wh = build_xlsx_by_warehouse(xlsx_data)

    print("🔍 수량 비교 중...")
    results = compare(xls_data, xlsx_by_wh)

    # 콘솔 요약 출력
    print("\n" + "="*70)
    print(f"{'납품처':<22} {'일치':>5} {'불일치':>6} {'XLS만':>6} {'XLSX만':>7} {'일치율':>7}")
    print("-"*70)
    total_m = total_mm = total_xo = total_xlo = 0
    for wh, res in results.items():
        m  = len(res["match"])
        mm = len(res["mismatch"])
        xo = len(res["xls_only"])
        xlo= len(res["xlsx_only"])
        t  = m + mm + xo + xlo
        r  = f"{round(m/t*100,1)}%" if t else "-"
        total_m += m; total_mm += mm; total_xo += xo; total_xlo += xlo
        print(f"  {wh:<20} {m:>5} {mm:>6} {xo:>6} {xlo:>7} {r:>7}")
    t = total_m + total_mm + total_xo + total_xlo
    overall = f"{round(total_m/t*100,1)}%" if t else "-"
    print("-"*70)
    print(f"  {'합계':<20} {total_m:>5} {total_mm:>6} {total_xo:>6} {total_xlo:>7} {overall:>7}")
    print("="*70)

    write_excel(results, xls_data, out)


if __name__ == "__main__":
    main()
