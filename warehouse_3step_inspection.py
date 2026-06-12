#!/usr/bin/env python3
"""
창고이동 3단계 검수 자동화
────────────────────────────────────────────────────────────────
Step 1. 이동처리 파일(F열) vs 라벨발행 파일(I열) 출고수량 비교
         · 이동처리: I열(비고)에 'TCT 케이터링 이동처리' 포함 행만 사용
         · 라벨발행: 취소/변경 행 자동 제외
         · 매칭 키: ERP코드 (날짜 무관, 전체 합산)

Step 2. 라벨발행 파일(L열 급품목코드) vs 작업내역 파일(G열 제품코드) 수량 비교
         · 작업내역 K열 수량 ÷ 2 (창고이동 2회)
         · 매칭 키: 급품목코드
         · 선작업 파일 첨부 시: 선작업 품목은 작업내역에 없는 것이 정상
           → "2단계_선작업(정상)" 시트로 분리하여 표시

사용법:
  python3 warehouse_3step_inspection.py <이동처리.xls> <라벨발행.xlsx> <작업내역.xlsx> [선작업.xlsx] [출력.xlsx]
────────────────────────────────────────────────────────────────
"""

import sys, os, re
from collections import defaultdict

try:
    import xlrd, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install xlrd openpyxl -q")
    import xlrd, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── 스타일 ───────────────────────────────────────────────────────
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def hdr(cell, bg="1F4E79"):
    cell.font  = Font(color="FFFFFF", bold=True, size=10)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = CENTER
    cell.border = BORDER

def dat(cell, align=LEFT, bold=False, bg=None):
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = align
    cell.border    = BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def clean_code(v) -> str:
    return re.sub(u'[\u200b\u200c\u200d\ufeff\u00a0]', '', str(v or '')).strip()


# ── 이동처리 파일 로드 ────────────────────────────────────────────
def load_warehouse(path):
    try:
        wb = xlrd.open_workbook(path)
    except Exception as e:
        print(f"[오류] 이동처리 파일: {e}"); sys.exit(1)
    ws = wb.sheet_by_index(0)
    TARGET = "TCT 케이터링 이동처리"

    qty_map   = defaultdict(float)
    info_map  = {}
    dates_map = defaultdict(set)

    for i in range(1, ws.nrows):
        note = str(ws.cell_value(i, 8))
        if TARGET not in note:
            continue
        m    = re.search(r'\((\d{4}-\d{2}-\d{2})\)', note)
        date = m.group(1) if m else str(ws.cell_value(i, 3)).strip()
        erp  = str(ws.cell_value(i, 9)).strip()
        name = str(ws.cell_value(i, 10)).strip()
        qty  = float(ws.cell_value(i, 5) or 0)

        qty_map[erp] += qty
        dates_map[erp].add(date)
        if erp not in info_map:
            info_map[erp] = {
                "품목명": name,
                "규격":   str(ws.cell_value(i, 11)).strip(),
                "단위":   str(ws.cell_value(i, 12)).strip(),
            }

    for erp in info_map:
        info_map[erp]["이동날짜"] = ", ".join(sorted(dates_map[erp]))

    return qty_map, info_map


# ── 라벨발행 파일 로드 ────────────────────────────────────────────
def load_label(path):
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"[오류] 라벨발행 파일: {e}"); sys.exit(1)
    ws = wb.active

    s1_qty  = defaultdict(float)
    s1_info = {}
    s2_qty  = defaultdict(float)
    s2_info = {}
    canceled = 0

    for i in range(2, ws.max_row + 1):
        state = ws.cell(i, 16).value
        if state in ('취소', '변경'):
            canceled += 1
            continue
        제품명   = str(ws.cell(i, 7).value  or '').strip()
        출고수량 = float(ws.cell(i, 9).value or 0)
        단위     = str(ws.cell(i, 10).value or '').strip()
        규격     = str(ws.cell(i, 11).value or '').strip()
        급코드   = clean_code(ws.cell(i, 12).value)
        erp코드  = str(ws.cell(i, 13).value or '').strip()

        s1_qty[erp코드] += 출고수량
        if erp코드 not in s1_info:
            s1_info[erp코드] = {
                "ERP코드": erp코드, "급품목코드": 급코드,
                "제품명": 제품명, "규격": 규격, "단위": 단위,
            }

        s2_qty[급코드] += 출고수량
        if 급코드 not in s2_info:
            s2_info[급코드] = {
                "급품목코드": 급코드, "ERP코드": erp코드,
                "제품명": 제품명, "규격": 규격, "단위": 단위,
            }

    print(f"  라벨발행 취소/변경 제외: {canceled}건")
    return s1_qty, s1_info, s2_qty, s2_info


# ── 작업내역 파일 로드 ────────────────────────────────────────────
def load_catering(path):
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"[오류] 작업내역 파일: {e}"); sys.exit(1)
    ws = wb.active
    qty_map  = defaultdict(float)
    info_map = {}
    for i in range(4, ws.max_row + 1):
        d = ws.cell(i, 3).value
        if not d or str(d) == '센터명':
            continue
        code = clean_code(ws.cell(i, 7).value)
        name = str(ws.cell(i, 8).value or '').strip()
        qty  = float(ws.cell(i, 11).value or 0)
        qty_map[code] += qty
        if code not in info_map:
            info_map[code] = {
                "제품코드": code, "제품명": name,
                "규격": str(ws.cell(i, 9).value or '').strip(),
                "단위": str(ws.cell(i, 10).value or '').strip(),
            }
    qty_half = {k: v / 2 for k, v in qty_map.items()}
    return qty_half, info_map


# ── 선작업 파일 로드 (선택) ───────────────────────────────────────
def load_prework(path):
    """
    선작업-pre_qty 시트에서 품목코드 목록 수집.
    해당 코드는 작업내역 파일에 없는 것이 정상 (선처리 완료).
    반환: {품목코드 -> {품목명, 급식사, 선작업qty합, 보정후합}}
    """
    if not path:
        return {}
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"[경고] 선작업 파일 로드 실패 (무시): {e}")
        return {}

    # '선작업-pre_qty' 시트 우선, 없으면 첫 번째 시트
    ws = wb['선작업-pre_qty'] if '선작업-pre_qty' in wb.sheetnames else wb.active
    items = {}
    for i in range(2, ws.max_row + 1):
        급식사  = str(ws.cell(i, 1).value or '').strip()
        code   = clean_code(ws.cell(i, 2).value)
        name   = str(ws.cell(i, 3).value or '').strip()
        pre_q  = float(ws.cell(i, 7).value or 0)  # 선작업qty
        fix_q  = float(ws.cell(i, 8).value or 0)  # 보정후
        if not code:
            continue
        if code not in items:
            items[code] = {"품목명": name, "급식사": 급식사,
                           "선작업qty": 0.0, "보정후": 0.0}
        items[code]["선작업qty"] += pre_q
        items[code]["보정후"]    += fix_q

    print(f"  선작업 품목: {len(items)}종")
    return items


# ── 비교 함수 ─────────────────────────────────────────────────────
def compare(a_qty, a_info, b_qty, b_info, col_a, col_b):
    all_keys = set(a_qty) | set(b_qty)
    matched, qty_diff, a_only, b_only = [], [], [], []

    for code in sorted(all_keys):
        aq = a_qty.get(code)
        bq = b_qty.get(code)
        ai = a_info.get(code, {})
        bi = b_info.get(code, {})

        name = (ai.get("품목명") or ai.get("제품명") or
                bi.get("품목명") or bi.get("제품명") or "")
        row = {
            "코드":     code,
            "품목명":   name,
            "규격":     ai.get("규격") or bi.get("규격", ""),
            "단위":     ai.get("단위") or bi.get("단위", ""),
            "이동날짜": ai.get("이동날짜", ""),
            col_a:      aq,
            col_b:      bq,
            "차이":     (bq or 0) - (aq or 0),
        }
        if aq is not None and bq is not None:
            (matched if abs(aq - bq) < 0.001 else qty_diff).append(row)
        elif aq is not None:
            a_only.append(row)
        else:
            b_only.append(row)

    return matched, qty_diff, a_only, b_only


def split_prework(matched, qty_diff, lbl_only, cat_only, prework_codes):
    """
    Step2 비교 결과 전체에서 선작업 코드 행을 분리.
    선작업 코드는 어느 분류(일치/불일치/한쪽만)에 있든 별도 시트로.
    """
    def _split(rows):
        pre, normal = [], []
        for row in rows:
            if row["코드"] in prework_codes:
                r = dict(row)
                r["선작업qty"] = prework_codes[row["코드"]]["선작업qty"]
                r["보정후"]    = prework_codes[row["코드"]]["보정후"]
                r["급식사"]    = prework_codes[row["코드"]]["급식사"]
                pre.append(r)
            else:
                normal.append(row)
        return pre, normal

    pre_m, matched   = _split(matched)
    pre_d, qty_diff  = _split(qty_diff)
    pre_l, lbl_only  = _split(lbl_only)
    pre_c, cat_only  = _split(cat_only)
    pre_rows = pre_m + pre_d + pre_l + pre_c
    return pre_rows, matched, qty_diff, lbl_only, cat_only


# ── 시트 작성 ─────────────────────────────────────────────────────
def write_sheet(wb, title, rows, hdr_bg, row_bg, col_a, col_b,
                show_date=True, extra_cols=None):
    ws = wb.create_sheet(title)
    base_h = ["코드", "품목명", "규격", "단위"]
    base_k = ["코드", "품목명", "규격", "단위"]
    if show_date:
        base_h.append("이동날짜"); base_k.append("이동날짜")
    base_h += [col_a, col_b, "차이"]
    base_k += [col_a, col_b, "차이"]
    if extra_cols:
        for eh, ek in extra_cols:
            base_h.append(eh); base_k.append(ek)

    w_base = [16, 38, 18, 7]
    if show_date: w_base.append(22)
    w_base += [16, 16, 10]
    if extra_cols: w_base += [14] * len(extra_cols)

    for c, h in enumerate(base_h, 1):
        hdr(ws.cell(1, c, h), hdr_bg)

    for r, item in enumerate(rows, 2):
        bg = row_bg if r % 2 == 0 else None
        for c, key in enumerate(base_k, 1):
            val = item.get(key)
            if val is None: val = "-"
            cell = ws.cell(r, c, val)
            align = CENTER if key in ("코드","단위",col_a,col_b,"차이","이동날짜","선작업qty","보정후") else LEFT
            cell_bg = bg
            if key == "차이" and isinstance(val, (int, float)) and abs(val) > 0.001:
                cell_bg = "FFD7D7"
            dat(cell, align, bg=cell_bg)

    for c, w in enumerate(w_base, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(base_h))}1"


# ── xlsx 저장 ─────────────────────────────────────────────────────
def save_all(s1_matched, s1_diff, s1_wh_only, s1_lbl_only,
             s2_matched, s2_diff, s2_lbl_only, s2_cat_only,
             s2_pre_rows, has_prework, output_path):
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "검수요약"
    ws_sum.column_dimensions["A"].width = 46
    ws_sum.column_dimensions["B"].width = 16

    ws_sum.merge_cells("A1:B1")
    hdr(ws_sum.cell(1, 1, "창고이동 3단계 검수 결과"), "1F4E79")

    s1c = len(s1_matched) + len(s1_diff)
    s2c = len(s2_matched) + len(s2_diff)

    rows_data = [
        ("── STEP 1: 이동처리(F열) vs 라벨발행(I열) ──", ""),
        ("  전체 비교 항목 (ERP코드 기준)",
         len(s1_matched)+len(s1_diff)+len(s1_wh_only)+len(s1_lbl_only)),
        ("  ✅ 일치",                     len(s1_matched)),
        ("  ❌ 수량 불일치 (양쪽 존재)",   len(s1_diff)),
        ("  ⚠️  이동처리에만 있는 항목",   len(s1_wh_only)),
        ("  ⚠️  라벨발행에만 있는 항목",   len(s1_lbl_only)),
        ("  일치율 (공통항목 기준)",
         f"{len(s1_matched)/s1c*100:.1f}%" if s1c else "-"),
        ("", ""),
        ("── STEP 2: 라벨발행(L열) vs 작업내역(G열) ──", ""),
        ("  전체 비교 항목 (급품목코드 기준)",
         len(s2_matched)+len(s2_diff)+len(s2_lbl_only)+len(s2_cat_only)+len(s2_pre_rows)),
        ("  ✅ 일치",                      len(s2_matched)),
        ("  ❌ 수량 불일치 (양쪽 존재)",    len(s2_diff)),
        ("  ⚠️  라벨발행에만 있는 항목",    len(s2_lbl_only)),
        ("  ⚠️  작업내역에만 있는 항목",    len(s2_cat_only)),
    ]
    if has_prework:
        rows_data.append(
            ("  ✅ 선작업 항목 (작업내역 미존재 정상)", len(s2_pre_rows))
        )
    rows_data.append(
        ("  일치율 (공통항목 기준)",
         f"{len(s2_matched)/s2c*100:.1f}%" if s2c else "-")
    )

    for r, (label, val) in enumerate(rows_data, 2):
        bold = label.startswith("──")
        bg   = "D9E1F2" if bold else None
        dat(ws_sum.cell(r, 1, label), LEFT, bold=bold, bg=bg)
        dat(ws_sum.cell(r, 2, val if val != "" else None), CENTER, bold=bold, bg=bg)

    A, B = "이동처리(F열)", "라벨발행(I열)"
    C, D = "라벨발행(I열)", "작업내역(K÷2)"

    write_sheet(wb, "1단계_수량불일치",  s1_diff,     "C00000", "FCE4D6", A, B)
    write_sheet(wb, "1단계_이동처리만",  s1_wh_only,  "843C0C", "FFF2CC", A, B)
    write_sheet(wb, "1단계_라벨발행만",  s1_lbl_only, "7030A0", "EAD1F5", A, B, show_date=False)
    write_sheet(wb, "1단계_일치",        s1_matched,  "375623", "E2EFDA", A, B)
    write_sheet(wb, "2단계_수량불일치",  s2_diff,     "C00000", "FCE4D6", C, D, show_date=False)
    write_sheet(wb, "2단계_라벨발행만",  s2_lbl_only, "843C0C", "FFF2CC", C, D, show_date=False)
    write_sheet(wb, "2단계_작업내역만",  s2_cat_only, "7030A0", "EAD1F5", C, D, show_date=False)
    write_sheet(wb, "2단계_일치",        s2_matched,  "375623", "E2EFDA", C, D, show_date=False)
    if has_prework and s2_pre_rows:
        write_sheet(wb, "2단계_선작업(정상)", s2_pre_rows, "4472C4", "DAE8FC", C, D,
                    show_date=False,
                    extra_cols=[("선작업qty", "선작업qty"), ("보정후", "보정후"), ("급식사", "급식사")])

    wb.save(output_path)
    print(f"\n저장 완료: {output_path}")


# ── 메인 ─────────────────────────────────────────────────────────
def main():
    WH_PATH  = "/root/.claude/uploads/4b31195f-0362-5c7f-baf9-3f04e5c8214d/21498ca7-20260608133431.xls"
    LBL_PATH = "/root/.claude/uploads/4b31195f-0362-5c7f-baf9-3f04e5c8214d/cdbc6ee1-Alps_________260609_1144.xlsx"
    CAT_PATH = "/root/.claude/uploads/4b31195f-0362-5c7f-baf9-3f04e5c8214d/544784f8-_____20260608_1420.xlsx"
    PRE_PATH = "/root/.claude/uploads/4b31195f-0362-5c7f-baf9-3f04e5c8214d/cd9a4e9a-________20260605_1.xlsx"
    OUT_PATH = "창고이동_3단계검수결과.xlsx"

    if len(sys.argv) >= 4:
        WH_PATH, LBL_PATH, CAT_PATH = sys.argv[1], sys.argv[2], sys.argv[3]
    PRE_PATH = sys.argv[4] if len(sys.argv) >= 5 else PRE_PATH
    OUT_PATH = sys.argv[5] if len(sys.argv) >= 6 else OUT_PATH

    print("=" * 60)
    print("[STEP 1] 이동처리 파일 로드 (TCT 케이터링 이동처리만)...")
    wh_qty, wh_info = load_warehouse(WH_PATH)
    print(f"  항목 수: {len(wh_qty)}건 (ERP코드 기준 합산)")

    print("\n[STEP 1] 라벨발행 파일 로드 (취소/변경 자동 제외)...")
    lbl_s1_qty, lbl_s1_info, lbl_s2_qty, lbl_s2_info = load_label(LBL_PATH)
    print(f"  Step1 항목: {len(lbl_s1_qty)}건 / Step2 항목: {len(lbl_s2_qty)}건")

    print("\n[STEP 2] 작업내역 파일 로드 (수량 ÷2)...")
    cat_qty, cat_info = load_catering(CAT_PATH)
    print(f"  항목 수: {len(cat_qty)}건")

    print("\n[선작업 파일 로드]...")
    prework = load_prework(PRE_PATH)
    has_prework = bool(prework)

    # ── 비교
    s1_matched, s1_diff, s1_wh_only, s1_lbl_only = compare(
        wh_qty, wh_info, lbl_s1_qty, lbl_s1_info, "이동처리(F열)", "라벨발행(I열)")

    s2_matched_all, s2_diff_all, s2_lbl_only_all, s2_cat_only_all = compare(
        lbl_s2_qty, lbl_s2_info, cat_qty, cat_info, "라벨발행(I열)", "작업내역(K÷2)")

    # 선작업 코드: 어느 분류에 있든 별도 시트로 분리
    s2_pre_rows, s2_matched, s2_diff, s2_lbl_only, s2_cat_only = split_prework(
        s2_matched_all, s2_diff_all, s2_lbl_only_all, s2_cat_only_all, prework)

    s1c = len(s1_matched) + len(s1_diff)
    s2c = len(s2_matched) + len(s2_diff)

    print("\n" + "=" * 60)
    print("[ STEP 1: 이동처리(F열) vs 라벨발행(I열) ]")
    if s1c:
        print(f"  ✅ 일치:        {len(s1_matched):>5}건  ({len(s1_matched)/s1c*100:.1f}%)")
        print(f"  ❌ 수량불일치:  {len(s1_diff):>5}건  ({len(s1_diff)/s1c*100:.1f}%)")
    print(f"  ⚠️  이동처리만: {len(s1_wh_only):>5}건")
    print(f"  ⚠️  라벨발행만: {len(s1_lbl_only):>5}건")

    print("\n[ STEP 2: 라벨발행(L열) vs 작업내역(G열) ]")
    if s2c:
        print(f"  ✅ 일치:        {len(s2_matched):>5}건  ({len(s2_matched)/s2c*100:.1f}%)")
        print(f"  ❌ 수량불일치:  {len(s2_diff):>5}건  ({len(s2_diff)/s2c*100:.1f}%)")
    print(f"  ⚠️  라벨발행만: {len(s2_lbl_only):>5}건  (미확인 불일치)")
    print(f"  ⚠️  작업내역만: {len(s2_cat_only):>5}건")
    if has_prework:
        print(f"  ✅ 선작업(정상): {len(s2_pre_rows):>4}건  (작업내역 미존재 정상)")
    print("=" * 60)

    save_all(s1_matched, s1_diff, s1_wh_only, s1_lbl_only,
             s2_matched, s2_diff, s2_lbl_only, s2_cat_only,
             s2_pre_rows, has_prework, OUT_PATH)


if __name__ == "__main__":
    main()
