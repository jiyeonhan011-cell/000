#!/usr/bin/env python3
"""
창고이동 검수 - Streamlit 버전
실행: streamlit run app.py
"""

import os, re, sys, io, tempfile
from pathlib import Path
from collections import defaultdict

try:
    import streamlit as st
    import xlrd, openpyxl
    from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install streamlit xlrd openpyxl -q")
    import streamlit as st
    import xlrd, openpyxl
    from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── 검수 로직 ──────────────────────────────────────────────────────

BORDER_STYLE = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def clean_code(v):
    return re.sub(u'[​‌‍﻿ ]', '', str(v or '')).strip()



# 파일 간 단위가 달라 수량이 다르게 보이는 품목 — 환산비 적용 후 일치하면 "단위다름(정상)"
# 값 형식: (환산배수, 품목명, 벤더사, 급식코드, ERP코드)
_UNIT_DIFF_DEFAULT = {
    "NA603095":    (168, "새찬 오이피클 일회용/중국산 80g (1BOX=168EA)",  "새찬",   "1000464464", "NA603095"),
    "1000464464":  (168, "새찬 오이피클 일회용/중국산 80g (1BOX=168EA)",  "새찬",   "1000464464", "NA603095"),
    "153325":      (168, "새찬 오이피클 일회용/중국산 80g (1BOX=168EA)",  "새찬",   "153325",      "NA603095"),
    "482644":      (168, "새찬 오이피클 일회용/중국산 80g (1BOX=168EA)",  "새찬",   "482644",      "NA603095"),
    "NB701107":    (30,  "청수식품 맑은물 냉면육수 300g/동치미맛 (1BOX=30EA)", "청수식품", "", "NB701107"),
    "NB706070":    (30,  "영미 동치미 냉면육수 310g (1BOX=30EA)",          "영미",   "", "NB706070"),
    "NE101062":    (6,   "삼원 튀김스프링 회오리감자 700g/냉동 (1BOX=6PAK)", "삼원",  "", "NE101062"),
    "NE101075":    (6,   "화영푸드 회오리감자 800g/냉동 (1BOX=6PAK)",      "화영푸드", "", "NE101075"),
    "NH101013":    (10,  "푸드림 미니바 슈가스틱 (1BOX=10EA)",             "푸드림", "", "NH101013"),
    "NH512021":    (10,  "식예원 가쓰오맛 후리가께 50g (1PAC=10EA)",       "식예원", "", "NH512021"),
    "NH610019":    (20,  "신진 신화당 50g (1PAK=20EA)",                    "신진",   "", "NH610019"),
    "NI307040":    (24,  "진로 하이트제로 350ml/무알콜 (1BOX=24EA)",       "진로",   "", "NI307040"),
    "NI309007":    (12,  "랭거스 망고쥬스 449ml (1BOX=12EA)",              "랭거스", "", "NI309007"),
    "NL102043":    (2,   "유진 카사바칩 1.2kg (1BOX=2EA)",                 "유진",   "", "NL102043"),
}

UNIT_FILE = Path(__file__).parent / "unit_config.json"

def load_unit_config():
    import json
    if UNIT_FILE.exists():
        try:
            data = json.loads(UNIT_FILE.read_text(encoding="utf-8"))
            result = {}
            for k, v in data.items():
                t = tuple(v)
                # 구형 형식 (factor, desc) → (factor, 품목명, 벤더사, 급식코드, ERP코드)
                if len(t) == 2:
                    t = (t[0], t[1], "", "", "")
                # 벤더사가 비어있으면 기본값에서 채움
                if len(t) > 2 and not t[2] and k in _UNIT_DIFF_DEFAULT:
                    t = (t[0], t[1], _UNIT_DIFF_DEFAULT[k][2], t[3], t[4])
                result[k] = t
            return result
        except: pass
    return dict(_UNIT_DIFF_DEFAULT)

def save_unit_config(data):
    import json
    UNIT_FILE.write_text(json.dumps({k: list(v) for k,v in data.items()}, ensure_ascii=False, indent=2), encoding="utf-8")

UNIT_DIFF_NORMAL = load_unit_config()

def _unit_diff_lookup(code, vendor=""):
    """벤더사별 등록이 있으면 우선 적용, 없으면 코드 공통 등록을 사용"""
    if vendor:
        info = UNIT_DIFF_NORMAL.get(f"{code}::{vendor}")
        if info: return info
    return UNIT_DIFF_NORMAL.get(code)

def _unit_diff_ok(row, col_a, col_b):
    """환산비 적용 후 두 수량이 일치하는지 확인"""
    info = _unit_diff_lookup(row["코드"], row.get("벤더사",""))
    if not info: return False
    factor = info[0]
    a = row.get(col_a) or 0
    b = row.get(col_b) or 0
    return abs(a * factor - b) < 0.001 or abs(b * factor - a) < 0.001

def load_warehouse(path):
    wb  = xlrd.open_workbook(path)
    ws  = wb.sheet_by_index(0)
    TARGET = "TCT 케이터링 이동처리"
    qty_map, info_map, dates_map = defaultdict(float), {}, defaultdict(set)
    for i in range(1, ws.nrows):
        note = str(ws.cell_value(i, 8))
        if TARGET not in note: continue
        m    = re.search(r'\((\d{4}-\d{2}-\d{2})\)', note)
        date = m.group(1) if m else str(ws.cell_value(i, 3)).strip()
        erp  = str(ws.cell_value(i, 9)).strip()
        qty  = float(ws.cell_value(i, 5) or 0)
        qty_map[erp] += qty
        dates_map[erp].add(date)
        if erp not in info_map:
            info_map[erp] = {
                "품목명": str(ws.cell_value(i,10)).strip(),
                "규격":   str(ws.cell_value(i,11)).strip(),
                "단위":   str(ws.cell_value(i,12)).strip(),
                "벤더사": str(ws.cell_value(i,21)).strip() if ws.ncols > 21 else "",
            }
    for erp in info_map:
        info_map[erp]["이동날짜"] = ", ".join(sorted(dates_map[erp]))
    return qty_map, info_map

BOX_UNITS = {'BOX','PAC','PAK','B','박스'}

def _to_ea(qty, 단위, code, vendor=""):
    """단위가 BOX계열이고 환산배수가 등록된 품목이면 EA로 환산 (벤더사별 등록 우선)"""
    info = _unit_diff_lookup(code, vendor)
    if 단위.upper() in BOX_UNITS and info:
        return qty * info[0]
    return qty

def load_label(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    s1_qty, s1_qty_cmp = defaultdict(float), defaultdict(float)
    s2_qty, s2_qty_cmp = defaultdict(float), defaultdict(float)
    s1_info, s2_info = {}, {}
    canceled = 0
    for i in range(2, ws.max_row + 1):
        state = ws.cell(i,16).value
        if state in ('취소','변경'):
            canceled += 1; continue
        제품명   = str(ws.cell(i,7).value  or '').strip()
        try:
            출고수량 = float(ws.cell(i,9).value or 0)
        except (ValueError, TypeError):
            continue
        단위     = str(ws.cell(i,10).value or '').strip()
        규격     = str(ws.cell(i,11).value or '').strip()
        급코드   = clean_code(ws.cell(i,12).value)
        erp코드  = str(ws.cell(i,13).value or '').strip()
        벤더사   = str(ws.cell(i,8).value  or '').strip()
        s1_qty[erp코드] += 출고수량                                       # 원본 (합계 표시용)
        s1_qty_cmp[erp코드] += _to_ea(출고수량, 단위, erp코드, 벤더사)  # 환산 (비교용)
        if erp코드 not in s1_info:
            s1_info[erp코드] = {"ERP코드":erp코드,"급품목코드":급코드,"제품명":제품명,"규격":규격,"단위":단위,"벤더사":벤더사}
        s2_qty[급코드] += 출고수량
        s2_qty_cmp[급코드] += _to_ea(출고수량, 단위, 급코드, 벤더사)
        if 급코드 not in s2_info:
            s2_info[급코드] = {"급품목코드":급코드,"ERP코드":erp코드,"제품명":제품명,"규격":규격,"단위":단위,"벤더사":벤더사}
    return s1_qty, s1_qty_cmp, s1_info, s2_qty, s2_qty_cmp, s2_info, canceled

def load_catering(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    qty_map, info_map = defaultdict(float), {}
    for i in range(4, ws.max_row + 1):
        d = ws.cell(i,3).value
        if not d or str(d)=='센터명': continue
        code = clean_code(ws.cell(i,7).value)
        qty  = float(ws.cell(i,11).value or 0)
        단위  = str(ws.cell(i,10).value or '').strip()
        qty_map[code] += qty
        if code not in info_map:
            info_map[code] = {
                "제품코드":code,"제품명":str(ws.cell(i,8).value or '').strip(),
                "규격":str(ws.cell(i,9).value or '').strip(),
                "단위":단위,
                "벤더사":str(ws.cell(i,5).value or '').strip(),
            }
    return qty_map, info_map

def apply_catering_div(qty_map, div2):
    return {k: v/2 for k,v in qty_map.items()} if div2 else dict(qty_map)

def load_prework(path):
    if not path or not os.path.exists(path): return {}, set()
    wb = openpyxl.load_workbook(path)
    # 선작업 코드
    items = {}
    if '선작업-pre_qty' in wb.sheetnames:
        ws = wb['선작업-pre_qty']
        for i in range(2, ws.max_row + 1):
            code = clean_code(ws.cell(i,2).value)
            if not code: continue
            if code not in items:
                items[code] = {"품목명":str(ws.cell(i,3).value or ''),"급식사":str(ws.cell(i,1).value or ''),"선작업qty":0.0,"보정후":0.0}
            items[code]["선작업qty"] += float(ws.cell(i,7).value or 0)
            items[code]["보정후"]    += float(ws.cell(i,8).value or 0)
    # BOX-EA 환산 코드
    box_ea = set()
    if 'BOX-EA환산' in wb.sheetnames:
        ws2 = wb['BOX-EA환산']
        for i in range(2, ws2.max_row + 1):
            code = clean_code(ws2.cell(i,2).value)
            if code: box_ea.add(code)
    return items, box_ea

def compare(a_qty, a_info, b_qty, b_info, col_a, col_b):
    matched, qty_diff, a_only, b_only = [], [], [], []
    for code in sorted(set(a_qty)|set(b_qty)):
        aq, bq = a_qty.get(code), b_qty.get(code)
        ai, bi = a_info.get(code,{}), b_info.get(code,{})
        name = ai.get("품목명") or ai.get("제품명") or bi.get("품목명") or bi.get("제품명") or ""
        row = {"코드":code,"품목명":name,
               "규격":ai.get("규격") or bi.get("규격",""),
               "단위":ai.get("단위") or bi.get("단위",""),
               "벤더사":ai.get("벤더사") or bi.get("벤더사",""),
               "이동날짜":ai.get("이동날짜",""),
               col_a:aq, col_b:bq, "차이":(bq or 0)-(aq or 0)}
        if aq is not None and bq is not None:
            (matched if abs(aq-bq)<0.001 else qty_diff).append(row)
        elif aq is not None: a_only.append(row)
        else: b_only.append(row)
    return matched, qty_diff, a_only, b_only

def split_prework(matched, qty_diff, lbl_only, cat_only, prework_codes, lbl_col):
    def _split(rows):
        pre, normal = [], []
        for row in rows:
            if row["코드"] in prework_codes:
                r = dict(row)
                r["선작업qty"] = prework_codes[row["코드"]]["선작업qty"]
                r["보정후"]    = prework_codes[row["코드"]]["보정후"]
                r["급식사"]    = prework_codes[row["코드"]]["급식사"]
                lbl_qty = row.get(lbl_col) or 0
                r["보정후차이"] = lbl_qty - r["보정후"]
                pre.append(r)
            else: normal.append(row)
        return pre, normal
    pm,m = _split(matched); pd,d = _split(qty_diff)
    pl,l = _split(lbl_only); pc,c = _split(cat_only)
    return pm+pd+pl+pc, m, d, l, c

def hdr_style(cell, bg="1F4E79"):
    cell.font  = XFont(color="FFFFFF", bold=True, size=10)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_STYLE

def dat_style(cell, center=False, bold=False, bg=None):
    cell.font      = XFont(bold=bold, size=10)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    cell.border    = BORDER_STYLE
    if bg: cell.fill = PatternFill("solid", fgColor=bg)

def write_xl_sheet(wb, title, rows, hdr_bg, row_bg, col_a, col_b, show_date=True, extra_cols=None):
    ws = wb.create_sheet(title)
    base_h = ["코드","품목명","규격","단위"]
    base_k = ["코드","품목명","규격","단위"]
    if show_date: base_h.append("이동날짜"); base_k.append("이동날짜")
    base_h += [col_a, col_b, "차이"]
    base_k += [col_a, col_b, "차이"]
    if extra_cols:
        for eh,ek in extra_cols: base_h.append(eh); base_k.append(ek)
    w_base = [16,38,18,7]
    if show_date: w_base.append(22)
    w_base += [16,16,10]
    if extra_cols: w_base += [14]*len(extra_cols)
    for c,h in enumerate(base_h,1): hdr_style(ws.cell(1,c,h), hdr_bg)
    for r,item in enumerate(rows,2):
        bg = row_bg if r%2==0 else None
        for c,key in enumerate(base_k,1):
            val = item.get(key); val = "-" if val is None else val
            cell = ws.cell(r,c,val)
            center = key in ("코드","단위",col_a,col_b,"차이","이동날짜","선작업qty","보정후","보정후차이")
            cell_bg = bg
            if key in ("차이","보정후차이") and isinstance(val,(int,float)) and abs(val)>0.001: cell_bg="FFD7D7"
            dat_style(cell, center=center, bg=cell_bg)
    for c,w in enumerate(w_base,1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(base_h))}1"


def run_inspection(wh_path, lbl_path, cat_path, pre_path, div2=True):
    wh_qty, wh_info = load_warehouse(wh_path)
    ls1q, ls1q_cmp, ls1i, ls2q, ls2q_cmp, ls2i, canceled = load_label(lbl_path)
    raw_cat_qty, cat_info = load_catering(cat_path)
    cat_qty = apply_catering_div(raw_cat_qty, div2)
    cat_col = "작업내역(K÷2)" if div2 else "작업내역(K)"
    prework, box_ea_codes = load_prework(pre_path) if pre_path else ({}, set())

    def is_box_ea(row):
        return row["코드"] in box_ea_codes or "혼용코드" in str(row.get("품목명",""))

    def restore_orig(rows, col_display, orig_qty, col_a, col_b):
        """비교는 환산값으로 했지만 표시는 원본값으로 복원"""
        for r in rows:
            orig = orig_qty.get(r["코드"])
            if orig is not None:
                r[col_display] = orig
                r["차이"] = (r.get(col_b) or 0) - (r.get(col_a) or 0)
        return rows

    s1m,s1d,s1w,s1l = compare(wh_qty,wh_info,ls1q_cmp,ls1i,"이동처리(F열)","라벨발행(I열)")
    for lst in (s1m, s1d, s1w, s1l):
        restore_orig(lst, "라벨발행(I열)", ls1q, "이동처리(F열)", "라벨발행(I열)")

    # BOX-EA 환산 품목을 수량불일치에서 분리 (파일 코드 + 품목명 혼용코드 자동감지)
    s1box = [r for r in s1d if is_box_ea(r)]
    s1d   = [r for r in s1d if not is_box_ea(r)]
    # 단위 다름: 환산비 적용 후 수량 일치하면 정상, 불일치하면 수량불일치로 유지
    s1unit = [r for r in s1d if _unit_diff_ok(r, "이동처리(F열)", "라벨발행(I열)")]
    s1d    = [r for r in s1d if not _unit_diff_ok(r, "이동처리(F열)", "라벨발행(I열)")]

    s2m_all,s2d_all,s2l_all,s2c_all = compare(ls2q,ls2i,cat_qty,cat_info,"라벨발행(I열)",cat_col)

    if prework:
        s2pre,s2m,s2d,s2l,s2c = split_prework(s2m_all,s2d_all,s2l_all,s2c_all,prework,"라벨발행(I열)")
    else:
        s2pre,s2m,s2d,s2l,s2c = [],s2m_all,s2d_all,s2l_all,s2c_all

    # BOX-EA 환산 품목을 2단계 수량불일치에서도 분리 (파일 코드 + 품목명 혼용코드 자동감지)
    s2box = [r for r in s2d if is_box_ea(r)]
    s2d   = [r for r in s2d if not is_box_ea(r)]
    # 단위 다름: 환산비 적용 후 수량 일치하면 정상, 불일치하면 수량불일치로 유지
    s2unit = [r for r in s2d if _unit_diff_ok(r, "라벨발행(I열)", "작업내역(K÷2)")]
    s2d    = [r for r in s2d if not _unit_diff_ok(r, "라벨발행(I열)", "작업내역(K÷2)")]

    wb = openpyxl.Workbook()
    ws_sum = wb.active; ws_sum.title = "검수요약"
    ws_sum.column_dimensions["A"].width=46; ws_sum.column_dimensions["B"].width=16
    ws_sum.merge_cells("A1:B1")
    hdr_style(ws_sum.cell(1,1,"창고이동 검수 결과"), "1F4E79")
    s1c = len(s1m)+len(s1d); s2c_cnt = len(s2m)+len(s2d)
    summary_rows = [
        ("── STEP 1: 이동처리(F열) vs 라벨발행(I열) ──",""),
        ("  전체 항목 (ERP코드 기준)", len(s1m)+len(s1d)+len(s1w)+len(s1l)+len(s1box)+len(s1unit)),
        ("  ✅ 일치", len(s1m)),("  ❌ 수량 불일치", len(s1d)),
        ("  ✅ BOX/EA환산(정상)", len(s1box)),
        ("  ✅ 단위다름(정상)", len(s1unit)),
        ("  ⚠️ 이동처리에만", len(s1w)),("  ⚠️ 라벨발행에만", len(s1l)),
        ("  일치율(공통기준)", f"{len(s1m)/s1c*100:.1f}%" if s1c else "-"),
        ("",""),
        ("── STEP 2: 라벨발행(L열) vs 작업내역(G열) ──",""),
        ("  전체 항목 (급품목코드 기준)", len(s2m)+len(s2d)+len(s2l)+len(s2c)+len(s2pre)+len(s2box)+len(s2unit)),
        ("  ✅ 일치", len(s2m)),("  ❌ 수량 불일치", len(s2d)),
        ("  ✅ BOX/EA환산(정상)", len(s2box)),
        ("  ✅ 단위다름(정상)", len(s2unit)),
        ("  ⚠️ 라벨발행에만", len(s2l)),("  ⚠️ 작업내역에만", len(s2c)),
        ("  ✅ 선작업(정상)", len(s2pre)),
        ("  일치율(공통기준)", f"{len(s2m)/s2c_cnt*100:.1f}%" if s2c_cnt else "-"),
    ]
    for r,(label,val) in enumerate(summary_rows,2):
        bold = label.startswith("──"); bg="D9E1F2" if bold else None
        dat_style(ws_sum.cell(r,1,label), bold=bold, bg=bg)
        dat_style(ws_sum.cell(r,2,val if val!="" else None), center=True, bold=bold, bg=bg)

    A,B = "이동처리(F열)","라벨발행(I열)"
    C,D = "라벨발행(I열)","작업내역(K÷2)"
    write_xl_sheet(wb,"1단계_수량불일치",s1d,"C00000","FCE4D6",A,B)
    write_xl_sheet(wb,"1단계_이동처리만",s1w,"843C0C","FFF2CC",A,B)
    write_xl_sheet(wb,"1단계_라벨발행만",s1l,"7030A0","EAD1F5",A,B,show_date=False)
    write_xl_sheet(wb,"1단계_일치",      s1m,"375623","E2EFDA",A,B)
    if s1box:
        write_xl_sheet(wb,"1단계_BOX_EA환산(정상)",s1box,"4472C4","DAE8FC",A,B)
    write_xl_sheet(wb,"2단계_수량불일치",s2d,"C00000","FCE4D6",C,D,show_date=False)
    write_xl_sheet(wb,"2단계_라벨발행만",s2l,"843C0C","FFF2CC",C,D,show_date=False)
    write_xl_sheet(wb,"2단계_작업내역만",s2c,"7030A0","EAD1F5",C,D,show_date=False)
    write_xl_sheet(wb,"2단계_일치",      s2m,"375623","E2EFDA",C,D,show_date=False)
    if s2pre:
        write_xl_sheet(wb,"2단계_선작업(정상)",s2pre,"4472C4","DAE8FC",C,D,show_date=False,
                       extra_cols=[("급식사","급식사"),("선작업qty","선작업qty"),("보정후","보정후"),("보정후차이","보정후차이")])
    if s2box:
        write_xl_sheet(wb,"2단계_BOX_EA환산(정상)",s2box,"4472C4","DAE8FC",C,D,show_date=False)
    if s1unit:
        write_xl_sheet(wb,"1단계_단위다름(정상)",s1unit,"6B6B6B","EFEFEF",A,B)
    if s2unit:
        write_xl_sheet(wb,"2단계_단위다름(정상)",s2unit,"6B6B6B","EFEFEF",C,D,show_date=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    A,B = "이동처리(F열)","라벨발행(I열)"
    C,D = "라벨발행(I열)",cat_col
    B_CMP = "라벨발행(환산)"
    D_CMP = "작업내역(환산)"

    def add_converted(rows, cmp_qty, orig_col, cmp_col_name, a_col):
        """환산수량 컬럼 추가 및 차이 재계산"""
        for r in rows:
            cmp = cmp_qty.get(r["코드"])
            if cmp is not None and abs(cmp - (r.get(orig_col) or 0)) > 0.001:
                r[cmp_col_name] = cmp
                r["차이"] = cmp - (r.get(a_col) or 0)
        return rows

    for lst in (s1m, s1d, s1w, s1l, s1box, s1unit):
        add_converted(lst, ls1q_cmp, B, B_CMP, A)
    for lst in (s2m, s2d, s2l, s2c, s2pre, s2box, s2unit):
        add_converted(lst, ls2q_cmp, D, D_CMP, C)

    def sumq(rows, col): return sum(r[col] for r in rows if isinstance(r.get(col), (int,float)))
    all_s1 = s1m+s1d+s1w+s1l+s1box+s1unit
    all_s2 = s2m+s2d+s2l+s2c+s2pre+s2box+s2unit

    return buf, {
        "s1_matched":len(s1m),"s1_diff":len(s1d),"s1_wh":len(s1w),"s1_lbl":len(s1l),
        "s1_box":len(s1box),"s1_unit":len(s1unit),
        "s1_rate":f"{len(s1m)/s1c*100:.1f}" if s1c else "0",
        "s1_wh_qty": sumq(all_s1, A), "s1_lbl_qty": sumq(all_s1, B),
        "s2_matched":len(s2m),"s2_diff":len(s2d),"s2_lbl":len(s2l),"s2_cat":len(s2c),
        "s2_pre":len(s2pre), "s2_box":len(s2box),"s2_unit":len(s2unit),
        "s2_rate":f"{len(s2m)/s2c_cnt*100:.1f}" if s2c_cnt else "0",
        "s2_lbl_qty": sumq(all_s2, C), "s2_cat_qty": sumq(all_s2, D),
        "canceled": canceled,
        "cols": {"A": A, "B": B, "B_CMP": B_CMP, "C": C, "D": D, "D_CMP": D_CMP},
        "rows": {
            "s1_diff": s1d, "s1_wh": s1w, "s1_lbl": s1l, "s1_box": s1box, "s1_matched": s1m,
            "s1_unit": s1unit,
            "s2_diff": s2d, "s2_lbl": s2l, "s2_cat": s2c, "s2_pre": s2pre, "s2_matched": s2m,
            "s2_box": s2box, "s2_unit": s2unit,
        },
    }


# ── Streamlit UI ───────────────────────────────────────────────────

st.set_page_config(
    page_title="창고이동 검수",
    page_icon="📦",
    layout="wide",
)

st.markdown("""
<style>
    /* 전체 배경 */
    .stApp { background-color: #f0f2f6; }

    /* 사이드바 */
    [data-testid="stSidebar"] {
        background-color: #1a2744;
    }
    [data-testid="stSidebar"] * { color: #e8eaf0 !important; }
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stCheckbox label,
    [data-testid="stSidebar"] .stTextInput label { color: #a0aec0 !important; font-size: 0.8rem !important; }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
        background-color: #253563 !important; border-color: #3a4f7a !important; color: #e8eaf0 !important;
    }
    [data-testid="stSidebar"] input {
        background-color: #253563 !important; border-color: #3a4f7a !important; color: #e8eaf0 !important;
    }
    [data-testid="stSidebar"] .stButton > button {
        background-color: #2563eb !important; color: white !important;
        border: none !important; border-radius: 6px !important;
        font-weight: 600 !important; width: 100%;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background-color: #1d4ed8 !important;
    }
    [data-testid="stSidebar"] hr { border-color: #2d3f6b !important; }

    /* 메트릭 카드 */
    div[data-testid="metric-container"] {
        background: white; border-radius: 10px; padding: 16px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.08);
        border-left: 4px solid #2563eb;
    }

    /* 헤더 */
    .wms-header {
        background: linear-gradient(135deg, #1a2744 0%, #2563eb 100%);
        padding: 20px 28px; border-radius: 12px; margin-bottom: 20px;
        display: flex; align-items: center; gap: 16px;
    }
    .wms-header h1 { color: white !important; margin: 0; font-size: 1.6rem; }
    .wms-header p  { color: #93c5fd !important; margin: 0; font-size: 0.9rem; }

    /* 섹션 헤더 */
    .section-header {
        background: white; border-radius: 10px; padding: 12px 20px;
        margin: 16px 0 8px 0; border-left: 5px solid #2563eb;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
        font-weight: 700; font-size: 1rem; color: #1a2744;
    }

    /* 상태 배너 */
    .status-ok  { background:#dcfce7; border:1.5px solid #16a34a; border-radius:10px; padding:16px 20px; color:#15803d; font-weight:700; font-size:1.05rem; }
    .status-err { background:#fee2e2; border:1.5px solid #dc2626; border-radius:10px; padding:16px 20px; color:#b91c1c; font-weight:700; font-size:1.05rem; }

    /* expander */
    [data-testid="stExpander"] { background: white; border-radius: 10px; border: 1px solid #e2e8f0 !important; }

    /* 구분선 */
    hr { border-color: #e2e8f0 !important; }
</style>
""", unsafe_allow_html=True)

# ── 헬퍼 함수 ──────────────────────────────────────────────────────

def list_files(folder, prefix):
    p = Path(folder)
    if not p.is_dir():
        return []
    files = sorted(
        [f for f in p.iterdir() if f.name.startswith(prefix) and f.suffix.lower() in ('.xls','.xlsx')],
        key=lambda f: f.stat().st_mtime, reverse=True
    )
    return files

def find_latest(folder, prefix):
    files = list_files(folder, prefix)
    return str(files[0]) if files else None

def file_selectbox(label, folder, prefix, container):
    files = list_files(folder, prefix)
    if not files:
        container.error(f"{prefix} 파일 없음")
        return None
    names = [f.name for f in files]
    chosen = container.selectbox(label, names, index=0)
    return str(Path(folder) / chosen)

CONFIG_FILE = Path(__file__).parent / "config.json"

def load_config():
    if CONFIG_FILE.exists():
        import json
        try: return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except: pass
    return {}

def save_config(data):
    import json
    CONFIG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# ── 사이드바 ───────────────────────────────────────────────────────

cfg = load_config()
DEFAULT_FOLDER = str(Path(__file__).parent / "검수파일")
saved_path = cfg.get("folder_path", DEFAULT_FOLDER)

for sub in ("이동처리", "라벨발행", "작업내역", "선작업"):
    (Path(saved_path) / sub).mkdir(parents=True, exist_ok=True)
if "folder_path" not in cfg:
    save_config({**cfg, "folder_path": DEFAULT_FOLDER})

with st.sidebar:
    st.markdown("## 📦 창고이동 검수")
    st.markdown("---")
    page = st.radio("메뉴", ["🔍 검수", "⚙️ 환산비 관리"], label_visibility="collapsed")
    st.markdown("---")

    st.markdown("##### 📁 검수 파일 폴더")
    folder_path = st.text_input("폴더 경로", value=saved_path, label_visibility="collapsed")
    if folder_path and folder_path != saved_path:
        for sub in ("이동처리", "라벨발행", "작업내역", "선작업"):
            (Path(folder_path) / sub).mkdir(parents=True, exist_ok=True)
        save_config({**cfg, "folder_path": folder_path})
        saved_path = folder_path
    folder_path = folder_path or saved_path

    st.markdown("---")
    st.markdown("##### 📄 파일 선택")

    wh_path = lbl_path = cat_path = pre_path = None
    if folder_path:
        wh_path  = file_selectbox("① 이동처리",  Path(folder_path) / "이동처리",  "이동처리",  st.sidebar)
        lbl_path = file_selectbox("② 라벨발행",  Path(folder_path) / "라벨발행",  "라벨발행",  st.sidebar)
        cat_path = file_selectbox("③ 작업내역",  Path(folder_path) / "작업내역",  "작업내역",  st.sidebar)
        pre_files = list_files(Path(folder_path) / "선작업", "선작업")
        if pre_files:
            pre_names = ["(없음)"] + [f.name for f in pre_files]
            chosen = st.selectbox("④ 선작업 (선택)", pre_names, index=1)
            pre_path = str(Path(folder_path) / "선작업" / chosen) if chosen != "(없음)" else None
        else:
            st.info("선작업 파일 없음")

    st.markdown("---")
    st.markdown("##### ⚙️ 검수 옵션")
    div2 = st.checkbox("작업내역 ÷2 적용", value=True, help="2번 이동 처리된 경우 체크, 1번이면 해제")

    st.markdown("---")
    ready = bool(wh_path and lbl_path and cat_path)
    run_btn = st.button("🔍 검수 시작", disabled=not ready, use_container_width=True)
    if not ready:
        st.caption("이동처리 · 라벨발행 · 작업내역 파일이 필요합니다.")

# ── 메인 화면 ──────────────────────────────────────────────────────

st.markdown("""
<div class="wms-header">
  <div>
    <h1>📦 창고이동 검수 시스템</h1>
    <p>이동처리 → 라벨발행 → 작업내역 자동 비교</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ── 환산비 관리 페이지 ────────────────────────────────────────────
if page == "⚙️ 환산비 관리":
    st.markdown('<div class="section-header">⚙️ 단위 환산비 관리</div>', unsafe_allow_html=True)
    st.caption("파일마다 단위가 달라 수량이 다르게 보이는 품목을 등록합니다. 등록된 품목은 환산비 적용 후 일치하면 '단위다름(정상)'으로 분류됩니다.")
    st.markdown("")

    unit_data = load_unit_config()

    VENDOR_LIST = [
        "동원홈푸드", "딜리버리랩", "신세계", "아모제", "아워홈",
        "웰스토리", "푸드머스", "프레시웨이", "한화 푸디스트", "허브메카", "현대 그린푸드", "SPC",
    ]
    with st.expander("➕ 새 품목 추가", expanded=True):
        st.caption("같은 품목이라도 급식사마다 출고 단위(EA/BOX/PAK)가 다르면, 급식사를 지정해서 등록하세요. '전체 공통'은 급식사 구분 없이 적용됩니다.")
        r1c1, r1c2, r1c3, r1c4, r1c5, r1c6 = st.columns([2, 2, 2, 3, 1, 1])
        vendor_options = ["전체 공통"] + VENDOR_LIST + ["직접 입력"]
        vendor_choice = r1c1.selectbox("적용 급식사", options=vendor_options)
        if vendor_choice == "직접 입력":
            new_vendor = r1c1.text_input("급식사 입력", placeholder="예: 새찬", label_visibility="collapsed")
        elif vendor_choice == "전체 공통":
            new_vendor = ""
        else:
            new_vendor = vendor_choice
        new_급식   = r1c2.text_input("급식코드", placeholder="예: 1000464464")
        new_erp    = r1c3.text_input("ERP코드",  placeholder="예: NA603095")
        new_name   = r1c4.text_input("품목명",   placeholder="예: 새찬 오이피클 80g (1BOX=168EA)")
        new_factor = r1c5.number_input("환산배수", min_value=1, value=1, step=1)
        r1c6.markdown("<br>", unsafe_allow_html=True)
        add_btn = r1c6.button("추가", use_container_width=True)
        if add_btn:
            급식_key = new_급식.strip()
            erp_key  = new_erp.strip()
            if not 급식_key and not erp_key:
                st.error("급식코드 또는 ERP코드 중 하나 이상 입력해주세요.")
            else:
                added = []
                val = (int(new_factor), new_name.strip(), new_vendor.strip(), 급식_key, erp_key)
                for code in [erp_key, 급식_key]:
                    if not code: continue
                    store_key = f"{code}::{new_vendor.strip()}" if new_vendor.strip() else code
                    if store_key in unit_data:
                        st.warning(f"이미 등록된 코드입니다: {code}" + (f" ({new_vendor})" if new_vendor else ""))
                    else:
                        unit_data[store_key] = val
                        added.append(code)
                if added:
                    save_unit_config(unit_data)
                    UNIT_DIFF_NORMAL.update(unit_data)
                    st.success(f"추가됨: {', '.join(added)}")
                    st.rerun()

    st.markdown("")
    st.markdown(f"##### 등록된 품목 ({len(unit_data)}건)")

    # 헤더
    hc1,hc2,hc3,hc4,hc5,hc6 = st.columns([2,2,2,3,1,1])
    hc1.markdown("**적용 급식사**"); hc2.markdown("**급식코드**"); hc3.markdown("**ERP코드**")
    hc4.markdown("**품목명**"); hc5.markdown("**환산배수**"); hc6.markdown("")
    st.divider()

    for code, info in list(unit_data.items()):
        factor  = info[0]
        name    = info[1] if len(info) > 1 else ""
        vendor  = info[2] if len(info) > 2 else ""
        급식코드 = info[3] if len(info) > 3 else ""
        erp코드  = info[4] if len(info) > 4 else ""
        c1,c2,c3,c4,c5,c6 = st.columns([2,2,2,3,1,1])
        c1.write(vendor or "전체 공통")
        c2.code(급식코드 or "-")
        c3.code(erp코드 or "-")
        c4.write(name or "-")
        c5.write(f"×{factor}")
        del_btn = c6.button("삭제", key=f"del_{code}")
        if del_btn:
            del unit_data[code]
            save_unit_config(unit_data)
            UNIT_DIFF_NORMAL.clear()
            UNIT_DIFF_NORMAL.update(unit_data)
            st.success(f"삭제됨: {code}")
            st.rerun()

    st.stop()

# ── 검수 페이지 ───────────────────────────────────────────────────
if not run_btn:
    st.markdown("""
    <div style="background:white;border-radius:12px;padding:40px;text-align:center;color:#94a3b8;box-shadow:0 1px 4px rgba(0,0,0,0.06);">
        <div style="font-size:3rem;">📋</div>
        <div style="font-size:1.1rem;font-weight:600;margin-top:12px;color:#475569;">왼쪽 사이드바에서 파일을 선택하고 검수를 시작하세요</div>
    </div>
    """, unsafe_allow_html=True)

if run_btn:
    with st.spinner("검수 중..."):
        try:
            buf, result = run_inspection(wh_path, lbl_path, cat_path, pre_path, div2=div2)

            import pandas as pd

            def to_df(rows, cols, extra=None):
                if not rows: return None
                base = ["코드","품목명","규격","단위"]
                # 원본 + 환산 컬럼 순서 구성
                qty_keys = []
                for c in cols:
                    qty_keys.append(c)
                    cmp = c.replace("(I열)","(환산)").replace("(F열)","(환산)").replace("(K÷2)","(환산)").replace("(K)","(환산)")
                    if cmp != c and cmp in rows[0]:
                        qty_keys.append(cmp)
                all_keys = base + [k for k in qty_keys if k in rows[0]] + ["차이"]
                if extra:
                    all_keys += [k for k in extra if k in rows[0]]
                return pd.DataFrame([{k: r.get(k,"-") for k in all_keys} for r in rows])

            rows = result["rows"]
            A,B   = result["cols"]["A"], result["cols"]["B"]
            B_CMP = result["cols"]["B_CMP"]
            C,D   = result["cols"]["C"], result["cols"]["D"]
            D_CMP = result["cols"]["D_CMP"]

            def vf(row_list): return row_list

            total_issue  = result["s1_wh"] + result["s1_lbl"] + result["s2_diff"] + result["s2_lbl"] + result["s2_cat"]
            total_normal = result["s1_matched"] + result["s1_box"] + result["s1_unit"] + result["s2_matched"] + result["s2_pre"] + result["s2_box"] + result["s2_unit"]

            # ── 상태 배너 ──
            if total_issue == 0:
                st.markdown('<div class="status-ok">✅ 정상 — 확인이 필요한 항목이 없습니다.</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="status-err">🚨 확인 필요 — {total_issue}건의 항목을 확인해야 합니다.</div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # ── 요약 지표 ──
            c1,c2,c3,c4,c5,c6 = st.columns(6)
            c1.metric("🚨 확인 필요", total_issue)
            c2.metric("✅ 정상 처리", total_normal)
            c3.metric("이동처리 수량", f"{result['s1_wh_qty']:,.0f}")
            c4.metric("라벨발행 수량", f"{result['s1_lbl_qty']:,.0f}", delta=f"{result['s1_lbl_qty']-result['s1_wh_qty']:+,.0f}")
            c5.metric("작업내역 수량", f"{result['s2_cat_qty']:,.0f}", delta=f"{result['s2_cat_qty']-result['s2_lbl_qty']:+,.0f}")
            c6.metric("취소/변경 제외", f"{result['canceled']}건")

            st.markdown("---")

            # ── 확인 필요 항목 ──
            issue_cnt = total_issue + result["s1_diff"]
            st.markdown(f'<div class="section-header">🚨 확인 필요 항목 ({issue_cnt}건)</div>', unsafe_allow_html=True)

            if rows["s1_diff"]:
                with st.expander(f"이동처리 ↔ 라벨발행 수량불일치 ({result['s1_diff']}건)", expanded=True):
                    st.dataframe(to_df(rows["s1_diff"], [A,B]), use_container_width=True, hide_index=True)
            if rows["s1_wh"]:
                with st.expander(f"이동처리에만 있는 품목 ({result['s1_wh']}건)", expanded=True):
                    st.dataframe(to_df(rows["s1_wh"], [A,B]), use_container_width=True, hide_index=True)
            if rows["s1_lbl"]:
                with st.expander(f"라벨발행에만 있는 품목 ({result['s1_lbl']}건)", expanded=True):
                    st.dataframe(to_df(rows["s1_lbl"], [A,B]), use_container_width=True, hide_index=True)
            if rows["s2_diff"]:
                with st.expander(f"라벨발행 ↔ 작업내역 수량불일치 ({result['s2_diff']}건)", expanded=True):
                    st.dataframe(to_df(rows["s2_diff"], [C,D]), use_container_width=True, hide_index=True)
            if rows["s2_lbl"]:
                with st.expander(f"라벨발행에만 있는 품목 ({result['s2_lbl']}건)", expanded=True):
                    st.dataframe(to_df(rows["s2_lbl"], [C,D]), use_container_width=True, hide_index=True)
            if rows["s2_cat"]:
                with st.expander(f"작업내역에만 있는 품목 ({result['s2_cat']}건)", expanded=True):
                    st.dataframe(to_df(rows["s2_cat"], [C,D]), use_container_width=True, hide_index=True)
            if issue_cnt == 0:
                st.success("확인이 필요한 항목이 없습니다.")

            st.markdown("---")

            # ── 정상 처리 항목 ──
            st.markdown(f'<div class="section-header">✅ 정상 처리 항목 ({total_normal}건)</div>', unsafe_allow_html=True)

            if rows["s1_matched"]:
                with st.expander(f"이동처리 ↔ 라벨발행 일치 ({result['s1_matched']}건)"):
                    st.dataframe(to_df(rows["s1_matched"], [A,B]), use_container_width=True, hide_index=True)
            if rows["s1_unit"]:
                with st.expander(f"단위다름(정상) — 이동처리 ↔ 라벨발행 ({result['s1_unit']}건)"):
                    st.caption("원본 수량이 달라 보이지만 환산 후 일치하는 품목입니다.")
                    st.dataframe(to_df(rows["s1_unit"], [A,B]), use_container_width=True, hide_index=True)
            if rows["s1_box"]:
                s1box_err = [r for r in rows["s1_box"] if abs(r.get("차이") or 0) > 0.001]
                with st.expander(f"BOX/EA환산(정상) — 이동처리 ↔ 라벨발행 ({result['s1_box']}건){' ⚠️ 수량 확인 필요 ' + str(len(s1box_err)) + '건' if s1box_err else ''}", expanded=bool(s1box_err)):
                    if s1box_err:
                        st.warning("아래 항목은 차이가 있습니다. 환산 오류 여부를 확인해주세요.")
                    st.dataframe(to_df(rows["s1_box"], [A,B]), use_container_width=True, hide_index=True)
            if rows["s2_matched"]:
                with st.expander(f"라벨발행 ↔ 작업내역 일치 ({result['s2_matched']}건)"):
                    st.dataframe(to_df(rows["s2_matched"], [C,D]), use_container_width=True, hide_index=True)
            if rows["s2_unit"]:
                with st.expander(f"단위다름(정상) — 라벨발행 ↔ 작업내역 ({result['s2_unit']}건)"):
                    st.caption("원본 수량이 달라 보이지만 환산 후 일치하는 품목입니다.")
                    st.dataframe(to_df(rows["s2_unit"], [C,D]), use_container_width=True, hide_index=True)
            if rows["s2_box"]:
                s2box_err = [r for r in rows["s2_box"] if abs(r.get("차이") or 0) > 0.001]
                with st.expander(f"BOX/EA환산(정상) — 라벨발행 ↔ 작업내역 ({result['s2_box']}건){' ⚠️ 수량 확인 필요 ' + str(len(s2box_err)) + '건' if s2box_err else ''}", expanded=bool(s2box_err)):
                    if s2box_err:
                        st.warning("아래 항목은 차이가 있습니다. 환산 오류 여부를 확인해주세요.")
                    st.dataframe(to_df(rows["s2_box"], [C,D]), use_container_width=True, hide_index=True)
            if rows["s2_pre"]:
                with st.expander(f"선작업(정상) ({result['s2_pre']}건)"):
                    st.caption("보정후차이 = 라벨발행(I열) - 보정후 (0이면 선작업 반영 시 정상)")
                    st.dataframe(to_df(rows["s2_pre"], [C,D], extra=["급식사","선작업qty","보정후","보정후차이"]),
                                 use_container_width=True, hide_index=True)

            st.markdown("---")
            st.download_button(
                label="📥 결과 엑셀 다운로드",
                data=buf,
                file_name="창고이동_검수결과.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        except Exception as e:
            import traceback
            st.error(f"오류가 발생했습니다:\n\n```\n{traceback.format_exc()}\n```")
