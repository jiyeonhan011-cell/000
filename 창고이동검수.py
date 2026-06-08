"""
창고이동검수 자동화 프로그램
- I열(비고)에서 'TCT 케이터링 이동처리(날짜별)' 데이터만 추출
- 이동수량을 1차/2차로 나누기 2하여 검수
- 결과를 Excel 파일로 출력
"""

import sys
import os
import xlrd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from collections import defaultdict


# ── 스타일 정의 ──────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
TITLE_FILL   = PatternFill("solid", fgColor="2E75B6")
SUB_FILL     = PatternFill("solid", fgColor="D6E4F0")
OK_FILL      = PatternFill("solid", fgColor="C6EFCE")
ERR_FILL     = PatternFill("solid", fgColor="FFC7CE")
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")

WHITE_FONT   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(name="맑은 고딕", bold=True, size=10)
NORM_FONT    = Font(name="맑은 고딕", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def load_catering_rows(xls_path: str) -> list[dict]:
    """XLS에서 TCT 케이터링 이동처리 행만 추출"""
    wb = xlrd.open_workbook(xls_path)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, j)) for j in range(ws.ncols)]

    rows = []
    for i in range(1, ws.nrows):
        note = str(ws.cell_value(i, 8))          # I열 = index 8
        if "케이터링 이동처리" in note:
            row = {headers[j]: ws.cell_value(i, j) for j in range(ws.ncols)}
            rows.append(row)
    return rows


def build_summary(rows: list[dict]) -> dict:
    """
    이동번호별·품목코드별로 수량 합산 후 1차/2차 분리
    구조: {이동번호: {품목코드: {품목명, 규격, 단위, 총수량, 1차, 2차, 입고창고, 출고창고, 비고, 이동일자}}}
    """
    # 이동번호 메타 (비고, 날짜, 입출고창고) 저장
    move_meta: dict[str, dict] = {}
    # 이동번호 → 품목코드 → 수량 누계
    data: dict[str, dict[str, dict]] = defaultdict(lambda: defaultdict(lambda: {
        "품목명": "", "규격": "", "단위": "",
        "총수량": 0.0, "입고창고": "", "출고창고": ""
    }))

    for r in rows:
        mn   = r["이동번호"]
        code = r["품목코드"]
        qty  = float(r["이동수량"] or 0)

        if mn not in move_meta:
            move_meta[mn] = {
                "비고":    r["비고"],
                "이동일자": r["이동일자"],
                "입고창고": r["입고창고"],
                "출고창고": r["출고창고"],
            }

        d = data[mn][code]
        d["품목명"]   = r["품목명"]
        d["규격"]     = r["규격"]
        d["단위"]     = r["단위"]
        d["총수량"]  += qty
        d["입고창고"] = r["입고창고"]
        d["출고창고"] = r["출고창고"]

    # 1차/2차 계산 (총수량 ÷ 2)
    result = {}
    for mn in sorted(data.keys()):
        items = []
        for code, d in sorted(data[mn].items()):
            total = d["총수량"]
            half  = total / 2

            # 정수 여부에 따라 표시 형식 결정
            def fmt(v):
                return int(v) if v == int(v) else round(v, 2)

            ok = (total % 2 == 0)   # 짝수면 정확히 나눠짐
            items.append({
                "품목코드": code,
                "품목명":   d["품목명"],
                "규격":     d["규격"],
                "단위":     d["단위"],
                "총수량":   fmt(total),
                "1차":      fmt(half),
                "2차":      fmt(half),
                "입고창고": d["입고창고"],
                "출고창고": d["출고창고"],
                "검수":     "✔ 정상" if ok else "⚠ 소수점",
            })
        result[mn] = {
            "meta":  move_meta[mn],
            "items": items,
        }
    return result


def write_excel(summary: dict, output_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 기본 시트 제거

    # ── 시트 1: 전체 요약 ──────────────────────────────────────
    ws_all = wb.create_sheet("전체요약")
    _write_summary_sheet(ws_all, summary)

    # ── 시트 2~N: 이동번호별 상세 ─────────────────────────────
    for mn, info in summary.items():
        date_str = info["meta"]["비고"].replace("TCT 케이터링 이동처리", "").strip("()")
        sheet_name = f"{date_str}_{info['meta']['입고창고']}"[:31]
        ws = wb.create_sheet(sheet_name)
        _write_detail_sheet(ws, mn, info)

    wb.save(output_path)
    print(f"\n✅ 결과 파일 저장: {output_path}")


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _cell(ws, row, col, value, font=None, fill=None, align=None, border=True, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = font  or NORM_FONT
    c.fill   = fill  or PatternFill()
    c.alignment = align or LEFT
    if border:
        c.border = thin_border()
    if num_format:
        c.number_format = num_format
    return c


def _write_summary_sheet(ws, summary: dict):
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # 타이틀
    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1, value="창고이동검수 요약표 (TCT 케이터링 이동처리)")
    c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=14)
    c.fill = HEADER_FILL
    c.alignment = CENTER

    # 헤더
    headers = ["이동번호", "비고(날짜)", "이동일자", "출고창고", "입고창고",
               "품목 수", "총수량 합계", "1차 합계", "2차 합계", "비고"]
    for col, h in enumerate(headers, 1):
        _cell(ws, 2, col, h, font=WHITE_FONT, fill=TITLE_FILL, align=CENTER)

    row = 3
    for mn, info in summary.items():
        meta  = info["meta"]
        items = info["items"]
        total_qty = sum(it["총수량"] for it in items)
        half_qty  = total_qty / 2

        def fmt(v):
            return int(v) if v == int(v) else round(v, 2)

        vals = [
            mn,
            meta["비고"],
            meta["이동일자"],
            meta["출고창고"],
            meta["입고창고"],
            len(items),
            fmt(total_qty),
            fmt(half_qty),
            fmt(half_qty),
            "정상" if all(it["검수"].startswith("✔") for it in items) else "⚠ 소수점 발생",
        ]
        fill = OK_FILL if vals[-1] == "정상" else ERR_FILL
        for col, v in enumerate(vals, 1):
            _cell(ws, row, col, v,
                  fill=fill,
                  align=CENTER if col in (1,6,7,8,9,10) else LEFT)
        row += 1

    _set_col_widths(ws, [22, 34, 14, 14, 20, 8, 12, 12, 12, 16])


def _write_detail_sheet(ws, mn: str, info: dict):
    meta  = info["meta"]
    items = info["items"]

    # 제목
    ws.merge_cells("A1:J1")
    title = f"[{meta['비고']}]  {meta['출고창고']} → {meta['입고창고']}  (이동번호: {mn})"
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=12)
    c.fill = HEADER_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # 이동일자 표시
    ws.merge_cells("A2:J2")
    c2 = ws.cell(row=2, column=1, value=f"이동일자: {meta['이동일자']}    총 {len(items)}품목")
    c2.font = BOLD_FONT
    c2.fill = SUB_FILL
    c2.alignment = LEFT
    ws.row_dimensions[2].height = 18

    # 컬럼 헤더
    col_headers = ["No", "품목코드", "품목명", "규격", "단위",
                   "원본수량(합계)", "1차 수량\n(÷2)", "2차 수량\n(÷2)", "1차+2차 검증", "검수결과"]
    ws.row_dimensions[3].height = 32
    for col, h in enumerate(col_headers, 1):
        _cell(ws, 3, col, h, font=WHITE_FONT, fill=TITLE_FILL, align=CENTER)

    for idx, it in enumerate(items, 1):
        row = idx + 3
        fill = OK_FILL if it["검수"].startswith("✔") else ERR_FILL
        alt  = ALT_FILL if idx % 2 == 0 else PatternFill()
        base_fill = fill if not it["검수"].startswith("✔") else alt

        total = it["총수량"]
        half1 = it["1차"]
        half2 = it["2차"]
        check_val = half1 + half2

        def fmt(v):
            return int(v) if isinstance(v, float) and v == int(v) else v

        vals = [
            idx,
            it["품목코드"],
            it["품목명"],
            it["규격"],
            it["단위"],
            fmt(total),
            fmt(half1),
            fmt(half2),
            fmt(check_val),
            it["검수"],
        ]
        for col, v in enumerate(vals, 1):
            cell_fill = fill if col in (9, 10) else base_fill
            _cell(ws, row, col, v,
                  fill=cell_fill,
                  align=CENTER if col in (1, 5, 6, 7, 8, 9, 10) else LEFT)

    # 합계 행
    total_row = len(items) + 4
    ws.row_dimensions[total_row].height = 20
    total_qty = sum(it["총수량"] for it in items)
    half_sum  = total_qty / 2

    def fmt(v):
        return int(v) if v == int(v) else round(v, 2)

    ws.merge_cells(f"A{total_row}:E{total_row}")
    _cell(ws, total_row, 1, "합 계", font=BOLD_FONT, fill=TITLE_FILL, align=CENTER)
    for col, v in enumerate([fmt(total_qty), fmt(half_sum), fmt(half_sum), fmt(total_qty), ""], 6):
        _cell(ws, total_row, col, v, font=BOLD_FONT, fill=TITLE_FILL, align=CENTER)

    _set_col_widths(ws, [5, 14, 36, 18, 6, 14, 12, 12, 12, 14])


def main():
    if len(sys.argv) < 2:
        # 기본 경로 사용 (같은 폴더에 xls 파일이 있을 경우)
        xls_files = [f for f in os.listdir(".") if f.endswith(".xls") or f.endswith(".xlsx")]
        if not xls_files:
            print("사용법: python3 창고이동검수.py <파일경로.xls>")
            sys.exit(1)
        xls_path = xls_files[0]
        print(f"파일 자동 감지: {xls_path}")
    else:
        xls_path = sys.argv[1]

    if not os.path.exists(xls_path):
        print(f"❌ 파일을 찾을 수 없습니다: {xls_path}")
        sys.exit(1)

    base_name = os.path.splitext(os.path.basename(xls_path))[0]
    output_path = os.path.join(os.path.dirname(xls_path) or ".", f"{base_name}_케이터링검수결과.xlsx")

    print(f"📂 파일 읽는 중: {xls_path}")
    rows = load_catering_rows(xls_path)
    print(f"   → TCT 케이터링 이동처리 행: {len(rows)}건")

    if not rows:
        print("⚠  해당 조건의 데이터가 없습니다.")
        sys.exit(0)

    summary = build_summary(rows)

    print(f"\n📊 이동번호별 요약:")
    for mn, info in summary.items():
        total = sum(it["총수량"] for it in info["items"])
        half  = total / 2
        print(f"   {mn}  |  {info['meta']['비고']}")
        print(f"     출고창고: {info['meta']['출고창고']}  →  입고창고: {info['meta']['입고창고']}")
        print(f"     품목 수: {len(info['items'])}   총수량: {int(total) if total==int(total) else total}")
        print(f"     1차 수량: {int(half) if half==int(half) else half}   2차 수량: {int(half) if half==int(half) else half}")

    write_excel(summary, output_path)


if __name__ == "__main__":
    main()
