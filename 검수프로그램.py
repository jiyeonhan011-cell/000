#!/usr/bin/env python3
"""
창고이동 3단계 검수 프로그램
"""

import sys, os, re, threading
from collections import defaultdict
from pathlib import Path

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QFileDialog, QTextEdit, QProgressBar,
    QGroupBox, QLineEdit, QFrame, QSizePolicy, QMessageBox,
    QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QMimeData
from PyQt5.QtGui import QFont, QColor, QDragEnterEvent, QDropEvent, QPalette

try:
    import xlrd, openpyxl
    from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install xlrd openpyxl -q")
    import xlrd, openpyxl
    from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ════════════════════════════════════════════════════════
#  검수 로직 (warehouse_3step_inspection.py 와 동일)
# ════════════════════════════════════════════════════════

BORDER_STYLE = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def clean_code(v):
    return re.sub(u'[​‌‍﻿ ]', '', str(v or '')).strip()

def load_warehouse(path):
    wb  = xlrd.open_workbook(path)
    ws  = wb.sheet_by_index(0)
    TARGET = "TCT 케이터링 이동처리"
    qty_map, info_map, dates_map = defaultdict(float), {}, defaultdict(set)
    for i in range(1, ws.nrows):
        note = str(ws.cell_value(i, 8))
        if TARGET not in note: continue
        m    = re.search(r'\((\d{4}-\d{2}-\d{2})\)', note)
        date = m.group(1) if m else str(ws.cell_value(i, 3)).strip()
        erp  = str(ws.cell_value(i, 9)).strip()
        qty  = float(ws.cell_value(i, 5) or 0)
        qty_map[erp] += qty
        dates_map[erp].add(date)
        if erp not in info_map:
            info_map[erp] = {
                "품목명": str(ws.cell_value(i,10)).strip(),
                "규격":   str(ws.cell_value(i,11)).strip(),
                "단위":   str(ws.cell_value(i,12)).strip(),
            }
    for erp in info_map:
        info_map[erp]["이동날짜"] = ", ".join(sorted(dates_map[erp]))
    return qty_map, info_map

def load_label(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    s1_qty, s1_info, s2_qty, s2_info = defaultdict(float), {}, defaultdict(float), {}
    canceled = 0
    for i in range(2, ws.max_row + 1):
        state = ws.cell(i,16).value
        if state in ('취소','변경'):
            canceled += 1; continue
        제품명   = str(ws.cell(i,7).value  or '').strip()
        출고수량 = float(ws.cell(i,9).value or 0)
        단위     = str(ws.cell(i,10).value or '').strip()
        규격     = str(ws.cell(i,11).value or '').strip()
        급코드   = clean_code(ws.cell(i,12).value)
        erp코드  = str(ws.cell(i,13).value or '').strip()
        s1_qty[erp코드] += 출고수량
        if erp코드 not in s1_info:
            s1_info[erp코드] = {"ERP코드":erp코드,"급품목코드":급코드,"제품명":제품명,"규격":규격,"단위":단위}
        s2_qty[급코드] += 출고수량
        if 급코드 not in s2_info:
            s2_info[급코드] = {"급품목코드":급코드,"ERP코드":erp코드,"제품명":제품명,"규격":규격,"단위":단위}
    return s1_qty, s1_info, s2_qty, s2_info, canceled

def load_catering(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    qty_map, info_map = defaultdict(float), {}
    for i in range(4, ws.max_row + 1):
        d = ws.cell(i,3).value
        if not d or str(d)=='센터명': continue
        code = clean_code(ws.cell(i,7).value)
        qty  = float(ws.cell(i,11).value or 0)
        qty_map[code] += qty
        if code not in info_map:
            info_map[code] = {
                "제품코드":code,"제품명":str(ws.cell(i,8).value or '').strip(),
                "규격":str(ws.cell(i,9).value or '').strip(),
                "단위":str(ws.cell(i,10).value or '').strip(),
            }
    return {k:v/2 for k,v in qty_map.items()}, info_map

def load_prework(path):
    if not path or not os.path.exists(path): return {}
    wb = openpyxl.load_workbook(path)
    ws = wb['선작업-pre_qty'] if '선작업-pre_qty' in wb.sheetnames else wb.active
    items = {}
    for i in range(2, ws.max_row + 1):
        code = clean_code(ws.cell(i,2).value)
        if not code: continue
        if code not in items:
            items[code] = {"품목명":str(ws.cell(i,3).value or ''),"급식사":str(ws.cell(i,1).value or ''),"선작업qty":0.0,"보정후":0.0}
        items[code]["선작업qty"] += float(ws.cell(i,7).value or 0)
        items[code]["보정후"]    += float(ws.cell(i,8).value or 0)
    return items

def compare(a_qty, a_info, b_qty, b_info, col_a, col_b):
    matched, qty_diff, a_only, b_only = [], [], [], []
    for code in sorted(set(a_qty)|set(b_qty)):
        aq, bq = a_qty.get(code), b_qty.get(code)
        ai, bi = a_info.get(code,{}), b_info.get(code,{})
        name = ai.get("품목명") or ai.get("제품명") or bi.get("품목명") or bi.get("제품명") or ""
        row = {"코드":code,"품목명":name,
               "규격":ai.get("규격") or bi.get("규격",""),
               "단위":ai.get("단위") or bi.get("단위",""),
               "이동날짜":ai.get("이동날짜",""),
               col_a:aq, col_b:bq, "차이":(bq or 0)-(aq or 0)}
        if aq is not None and bq is not None:
            (matched if abs(aq-bq)<0.001 else qty_diff).append(row)
        elif aq is not None: a_only.append(row)
        else: b_only.append(row)
    return matched, qty_diff, a_only, b_only

def split_prework(matched, qty_diff, lbl_only, cat_only, prework_codes):
    def _split(rows):
        pre, normal = [], []
        for row in rows:
            if row["코드"] in prework_codes:
                r = dict(row)
                r["선작업qty"] = prework_codes[row["코드"]]["선작업qty"]
                r["보정후"]    = prework_codes[row["코드"]]["보정후"]
                r["급식사"]    = prework_codes[row["코드"]]["급식사"]
                pre.append(r)
            else: normal.append(row)
        return pre, normal
    pm,m = _split(matched); pd,d = _split(qty_diff)
    pl,l = _split(lbl_only); pc,c = _split(cat_only)
    return pm+pd+pl+pc, m, d, l, c

def hdr_style(cell, bg="1F4E79"):
    cell.font  = XFont(color="FFFFFF", bold=True, size=10)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_STYLE

def dat_style(cell, center=False, bold=False, bg=None):
    cell.font      = XFont(bold=bold, size=10)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    cell.border    = BORDER_STYLE
    if bg: cell.fill = PatternFill("solid", fgColor=bg)

def write_xl_sheet(wb, title, rows, hdr_bg, row_bg, col_a, col_b, show_date=True, extra_cols=None):
    ws = wb.create_sheet(title)
    base_h = ["코드","품목명","규격","단위"]
    base_k = ["코드","품목명","규격","단위"]
    if show_date: base_h.append("이동날짜"); base_k.append("이동날짜")
    base_h += [col_a, col_b, "차이"]
    base_k += [col_a, col_b, "차이"]
    if extra_cols:
        for eh,ek in extra_cols: base_h.append(eh); base_k.append(ek)
    w_base = [16,38,18,7]
    if show_date: w_base.append(22)
    w_base += [16,16,10]
    if extra_cols: w_base += [14]*len(extra_cols)
    for c,h in enumerate(base_h,1): hdr_style(ws.cell(1,c,h), hdr_bg)
    for r,item in enumerate(rows,2):
        bg = row_bg if r%2==0 else None
        for c,key in enumerate(base_k,1):
            val = item.get(key); val = "-" if val is None else val
            cell = ws.cell(r,c,val)
            center = key in ("코드","단위",col_a,col_b,"차이","이동날짜","선작업qty","보정후")
            cell_bg = bg
            if key=="차이" and isinstance(val,(int,float)) and abs(val)>0.001: cell_bg="FFD7D7"
            dat_style(cell, center=center, bg=cell_bg)
    for c,w in enumerate(w_base,1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(base_h))}1"

def run_inspection(wh_path, lbl_path, cat_path, pre_path, out_path, log_fn):
    log_fn("이동처리 파일 로드 중...")
    wh_qty, wh_info = load_warehouse(wh_path)
    log_fn(f"  TCT 케이터링 이동처리 항목: {len(wh_qty)}건")

    log_fn("라벨발행 파일 로드 중 (취소/변경 제외)...")
    ls1q,ls1i,ls2q,ls2i,canceled = load_label(lbl_path)
    log_fn(f"  취소/변경 제외: {canceled}건 | Step1: {len(ls1q)}건 | Step2: {len(ls2q)}건")

    log_fn("작업내역 파일 로드 중 (÷2 적용)...")
    cat_qty, cat_info = load_catering(cat_path)
    log_fn(f"  항목 수: {len(cat_qty)}건")

    prework = {}
    if pre_path:
        log_fn("선작업 파일 로드 중...")
        prework = load_prework(pre_path)
        log_fn(f"  선작업 품목: {len(prework)}종")

    log_fn("비교 수행 중...")
    s1m,s1d,s1w,s1l = compare(wh_qty,wh_info,ls1q,ls1i,"이동처리(F열)","라벨발행(I열)")
    s2m_all,s2d_all,s2l_all,s2c_all = compare(ls2q,ls2i,cat_qty,cat_info,"라벨발행(I열)","작업내역(K÷2)")

    if prework:
        s2pre,s2m,s2d,s2l,s2c = split_prework(s2m_all,s2d_all,s2l_all,s2c_all,prework)
    else:
        s2pre,s2m,s2d,s2l,s2c = [],s2m_all,s2d_all,s2l_all,s2c_all

    log_fn("엑셀 저장 중...")
    wb = openpyxl.Workbook()
    ws_sum = wb.active; ws_sum.title = "검수요약"
    ws_sum.column_dimensions["A"].width=46; ws_sum.column_dimensions["B"].width=16
    ws_sum.merge_cells("A1:B1")
    hdr_style(ws_sum.cell(1,1,"창고이동 3단계 검수 결과"), "1F4E79")
    s1c_cnt = len(s1m)+len(s1d); s2c_cnt = len(s2m)+len(s2d)
    summary = [
        ("── STEP 1: 이동처리(F열) vs 라벨발행(I열) ──",""),
        ("  전체 항목 (ERP코드 기준)", len(s1m)+len(s1d)+len(s1w)+len(s1l)),
        ("  ✅ 일치", len(s1m)),
        ("  ❌ 수량 불일치", len(s1d)),
        ("  ⚠️  이동처리에만", len(s1w)),
        ("  ⚠️  라벨발행에만", len(s1l)),
        ("  일치율(공통기준)", f"{len(s1m)/s1c_cnt*100:.1f}%" if s1c_cnt else "-"),
        ("",""),
        ("── STEP 2: 라벨발행(L열) vs 작업내역(G열) ──",""),
        ("  전체 항목 (급품목코드 기준)", len(s2m)+len(s2d)+len(s2l)+len(s2c)+len(s2pre)),
        ("  ✅ 일치", len(s2m)),
        ("  ❌ 수량 불일치", len(s2d)),
        ("  ⚠️  라벨발행에만", len(s2l)),
        ("  ⚠️  작업내역에만", len(s2c)),
        ("  ✅ 선작업(정상)", len(s2pre)),
        ("  일치율(공통기준)", f"{len(s2m)/s2c_cnt*100:.1f}%" if s2c_cnt else "-"),
    ]
    for r,(label,val) in enumerate(summary,2):
        bold = label.startswith("──"); bg="D9E1F2" if bold else None
        dat_style(ws_sum.cell(r,1,label), bold=bold, bg=bg)
        dat_style(ws_sum.cell(r,2,val if val!="" else None), center=True, bold=bold, bg=bg)

    A,B = "이동처리(F열)","라벨발행(I열)"
    C,D = "라벨발행(I열)","작업내역(K÷2)"
    write_xl_sheet(wb,"1단계_수량불일치",s1d,"C00000","FCE4D6",A,B)
    write_xl_sheet(wb,"1단계_이동처리만",s1w,"843C0C","FFF2CC",A,B)
    write_xl_sheet(wb,"1단계_라벨발행만",s1l,"7030A0","EAD1F5",A,B,show_date=False)
    write_xl_sheet(wb,"1단계_일치",      s1m,"375623","E2EFDA",A,B)
    write_xl_sheet(wb,"2단계_수량불일치",s2d,"C00000","FCE4D6",C,D,show_date=False)
    write_xl_sheet(wb,"2단계_라벨발행만",s2l,"843C0C","FFF2CC",C,D,show_date=False)
    write_xl_sheet(wb,"2단계_작업내역만",s2c,"7030A0","EAD1F5",C,D,show_date=False)
    write_xl_sheet(wb,"2단계_일치",      s2m,"375623","E2EFDA",C,D,show_date=False)
    if s2pre:
        write_xl_sheet(wb,"2단계_선작업(정상)",s2pre,"4472C4","DAE8FC",C,D,show_date=False,
                       extra_cols=[("선작업qty","선작업qty"),("보정후","보정후"),("급식사","급식사")])
    wb.save(out_path)

    results = {
        "s1_matched":len(s1m),"s1_diff":len(s1d),"s1_wh":len(s1w),"s1_lbl":len(s1l),
        "s1_rate":f"{len(s1m)/s1c_cnt*100:.1f}%" if s1c_cnt else "-",
        "s2_matched":len(s2m),"s2_diff":len(s2d),"s2_lbl":len(s2l),"s2_cat":len(s2c),
        "s2_pre":len(s2pre),
        "s2_rate":f"{len(s2m)/s2c_cnt*100:.1f}%" if s2c_cnt else "-",
    }
    return results


# ════════════════════════════════════════════════════════
#  GUI
# ════════════════════════════════════════════════════════

ACCENT   = "#1F4E79"
SUCCESS  = "#375623"
DANGER   = "#C00000"
WARNING  = "#843C0C"
BG       = "#F4F6FA"
CARD_BG  = "#FFFFFF"
TEXT     = "#1A1A2E"


class DropLineEdit(QLineEdit):
    """파일을 드래그&드롭할 수 있는 경로 입력창"""
    def __init__(self, ext_filter, placeholder="파일을 선택하거나 드래그하세요", parent=None):
        super().__init__(parent)
        self.ext_filter = ext_filter
        self.setPlaceholderText(placeholder)
        self.setAcceptDrops(True)
        self.setReadOnly(True)
        self.setStyleSheet("""
            QLineEdit {
                border: 2px dashed #B0BEC5;
                border-radius: 8px;
                padding: 8px 12px;
                background: #FAFAFA;
                color: #37474F;
                font-size: 13px;
            }
            QLineEdit:focus { border-color: #1F4E79; background: #EEF4FB; }
        """)

    def dragEnterEvent(self, e: QDragEnterEvent):
        if e.mimeData().hasUrls(): e.acceptProposedAction()

    def dropEvent(self, e: QDropEvent):
        urls = e.mimeData().urls()
        if urls:
            path = urls[0].toLocalFile()
            if any(path.lower().endswith(x) for x in self.ext_filter):
                self.setText(path)
                self.setStyleSheet(self.styleSheet().replace("#FAFAFA","#E8F5E9").replace("dashed","solid"))


class WorkerThread(QThread):
    log_signal     = pyqtSignal(str)
    done_signal    = pyqtSignal(dict)
    error_signal   = pyqtSignal(str)

    def __init__(self, wh, lbl, cat, pre, out):
        super().__init__()
        self.wh, self.lbl, self.cat, self.pre, self.out = wh, lbl, cat, pre, out

    def run(self):
        try:
            result = run_inspection(self.wh, self.lbl, self.cat, self.pre, self.out,
                                    self.log_signal.emit)
            self.done_signal.emit(result)
        except Exception as e:
            import traceback
            self.error_signal.emit(traceback.format_exc())


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("창고이동 3단계 검수 프로그램")
        self.setMinimumSize(820, 720)
        self.setStyleSheet(f"QMainWindow {{ background: {BG}; }}")
        self._build_ui()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(14)

        # ── 헤더
        title = QLabel("창고이동 3단계 검수")
        title.setFont(QFont("Arial", 18, QFont.Bold))
        title.setStyleSheet(f"color: {ACCENT};")
        sub = QLabel("이동처리 → 라벨발행 → 작업내역 수량 자동 비교")
        sub.setStyleSheet("color: #607D8B; font-size: 13px;")
        root.addWidget(title)
        root.addWidget(sub)

        line = QFrame(); line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("color: #CFD8DC;")
        root.addWidget(line)

        # ── 파일 선택 카드
        file_card = self._card("파일 선택")
        fc_layout = file_card.layout()

        self.wh_edit  = self._file_row(fc_layout, "① 이동처리 파일 *",  [".xls",".xlsx"])
        self.lbl_edit = self._file_row(fc_layout, "② 라벨발행 파일 *",  [".xlsx"])
        self.cat_edit = self._file_row(fc_layout, "③ 작업내역 파일 *",  [".xlsx"])
        self.pre_edit = self._file_row(fc_layout, "④ 선작업 파일  (선택)", [".xlsx"], required=False)
        self.out_edit = self._file_row(fc_layout, "⑤ 결과 저장 경로",   [".xlsx"], save=True)
        self.out_edit.setText(str(Path.home() / "창고이동_3단계검수결과.xlsx"))
        root.addWidget(file_card)

        # ── 실행 버튼
        btn_row = QHBoxLayout()
        self.run_btn = QPushButton("  검수 시작")
        self.run_btn.setFont(QFont("Arial", 13, QFont.Bold))
        self.run_btn.setFixedHeight(48)
        self.run_btn.setStyleSheet(f"""
            QPushButton {{
                background: {ACCENT}; color: white;
                border-radius: 10px; padding: 0 32px;
            }}
            QPushButton:hover {{ background: #2563A8; }}
            QPushButton:disabled {{ background: #B0BEC5; }}
        """)
        self.run_btn.clicked.connect(self._run)
        btn_row.addStretch(); btn_row.addWidget(self.run_btn); btn_row.addStretch()
        root.addLayout(btn_row)

        # ── 진행 상태
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.setVisible(False)
        self.progress.setFixedHeight(6)
        self.progress.setStyleSheet(f"""
            QProgressBar {{ border: none; background: #E0E0E0; border-radius: 3px; }}
            QProgressBar::chunk {{ background: {ACCENT}; border-radius: 3px; }}
        """)
        root.addWidget(self.progress)

        # ── 탭 (로그 / 결과)
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #CFD8DC; border-radius: 8px; background: white; }
            QTabBar::tab { padding: 8px 20px; font-size: 13px; }
            QTabBar::tab:selected { color: #1F4E79; border-bottom: 3px solid #1F4E79; font-weight: bold; }
        """)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFont(QFont("Consolas", 11))
        self.log_box.setStyleSheet("background:#1E1E2E; color:#A9B7C6; border:none; padding:8px;")
        self.tabs.addTab(self.log_box, "로그")

        self.result_widget = QWidget()
        self.result_layout = QVBoxLayout(self.result_widget)
        self.result_layout.setContentsMargins(12,12,12,12)
        self.result_layout.setSpacing(10)
        self.tabs.addTab(self.result_widget, "결과 요약")

        root.addWidget(self.tabs)

    def _card(self, title):
        box = QGroupBox(title)
        box.setStyleSheet(f"""
            QGroupBox {{
                background: {CARD_BG}; border: 1px solid #CFD8DC;
                border-radius: 10px; margin-top: 8px; padding: 12px;
                font-size: 13px; font-weight: bold; color: {ACCENT};
            }}
            QGroupBox::title {{ subcontrol-origin: margin; left: 14px; }}
        """)
        layout = QVBoxLayout(box)
        layout.setSpacing(8)
        return box

    def _file_row(self, parent_layout, label, exts, required=True, save=False):
        row = QHBoxLayout()
        lbl = QLabel(label)
        lbl.setFixedWidth(200)
        lbl.setStyleSheet("font-size: 13px; color: #37474F;")
        edit = DropLineEdit(exts)
        edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        btn = QPushButton("찾기")
        btn.setFixedWidth(60)
        btn.setFixedHeight(36)
        btn.setStyleSheet(f"""
            QPushButton {{ background: {'#E3F2FD' if not save else '#FFF8E1'};
                color: {ACCENT}; border: 1px solid #B0BEC5; border-radius: 6px; font-size: 12px; }}
            QPushButton:hover {{ background: #BBDEFB; }}
        """)

        if save:
            btn.clicked.connect(lambda: self._pick_save(edit, exts))
        else:
            btn.clicked.connect(lambda: self._pick_open(edit, exts, required))

        row.addWidget(lbl)
        row.addWidget(edit)
        row.addWidget(btn)
        parent_layout.addLayout(row)
        return edit

    def _pick_open(self, edit, exts, required):
        filt = "Excel Files (*.xlsx *.xls)" if ".xls" in exts else "Excel Files (*.xlsx)"
        path, _ = QFileDialog.getOpenFileName(self, "파일 선택", "", filt)
        if path:
            edit.setText(path)
            edit.setStyleSheet(edit.styleSheet().replace("#FAFAFA","#E8F5E9").replace("dashed","solid"))

    def _pick_save(self, edit, exts):
        path, _ = QFileDialog.getSaveFileName(self, "저장 경로", edit.text(), "Excel Files (*.xlsx)")
        if path:
            if not path.endswith(".xlsx"): path += ".xlsx"
            edit.setText(path)

    def _run(self):
        wh  = self.wh_edit.text().strip()
        lbl = self.lbl_edit.text().strip()
        cat = self.cat_edit.text().strip()
        pre = self.pre_edit.text().strip() or None
        out = self.out_edit.text().strip()

        if not wh or not lbl or not cat:
            QMessageBox.warning(self, "파일 미선택", "①②③ 필수 파일을 모두 선택하세요.")
            return
        if not out:
            QMessageBox.warning(self, "저장 경로", "결과 파일 저장 경로를 입력하세요.")
            return

        self.log_box.clear()
        self._clear_result()
        self.run_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.tabs.setCurrentIndex(0)

        self.worker = WorkerThread(wh, lbl, cat, pre, out)
        self.worker.log_signal.connect(self._log)
        self.worker.done_signal.connect(self._on_done)
        self.worker.error_signal.connect(self._on_error)
        self.worker.start()

    def _log(self, msg):
        self.log_box.append(f"<span style='color:#A9B7C6'>▶ {msg}</span>")

    def _on_done(self, r):
        self.run_btn.setEnabled(True)
        self.progress.setVisible(False)
        self._log("─" * 50)
        self._log("완료!")
        self._show_result(r)
        self.tabs.setCurrentIndex(1)

    def _on_error(self, tb):
        self.run_btn.setEnabled(True)
        self.progress.setVisible(False)
        self._log(f"<span style='color:#FF5252'>오류 발생:\n{tb}</span>")
        QMessageBox.critical(self, "오류", tb[-500:])

    def _clear_result(self):
        while self.result_layout.count():
            item = self.result_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()

    def _show_result(self, r):
        self._clear_result()

        def section(title, rows, color):
            box = QGroupBox(title)
            box.setStyleSheet(f"""
                QGroupBox {{ background: white; border: 2px solid {color};
                    border-radius: 8px; margin-top: 8px; font-size: 13px;
                    font-weight: bold; color: {color}; padding: 10px; }}
                QGroupBox::title {{ subcontrol-origin: margin; left: 12px; }}
            """)
            lay = QVBoxLayout(box)
            for label, val, clr in rows:
                row = QHBoxLayout()
                l = QLabel(label); l.setStyleSheet("font-size:13px;")
                v = QLabel(str(val)); v.setFont(QFont("Arial",14,QFont.Bold))
                v.setStyleSheet(f"color:{clr}; font-size:15px;")
                row.addWidget(l); row.addStretch(); row.addWidget(v)
                lay.addLayout(row)
            return box

        self.result_layout.addWidget(section(
            "STEP 1 │ 이동처리(F열) vs 라벨발행(I열)",
            [
                ("✅  일치",         r["s1_matched"], SUCCESS),
                ("❌  수량 불일치",   r["s1_diff"],    DANGER),
                ("⚠️   이동처리에만", r["s1_wh"],      WARNING),
                ("⚠️   라벨발행에만", r["s1_lbl"],     WARNING),
                ("📊  일치율",        r["s1_rate"],    ACCENT),
            ], ACCENT
        ))

        self.result_layout.addWidget(section(
            "STEP 2 │ 라벨발행(L열) vs 작업내역(G열)",
            [
                ("✅  일치",               r["s2_matched"], SUCCESS),
                ("❌  수량 불일치",         r["s2_diff"],    DANGER),
                ("⚠️   라벨발행에만",       r["s2_lbl"],     WARNING),
                ("⚠️   작업내역에만",       r["s2_cat"],     WARNING),
                ("📋  선작업(정상 제외)",   r["s2_pre"],     "#4472C4"),
                ("📊  일치율",              r["s2_rate"],    ACCENT),
            ], "#4472C4"
        ))

        # 저장 경로 안내
        out_path = self.out_edit.text()
        info = QLabel(f"📁  결과 저장 완료: {out_path}")
        info.setStyleSheet("color:#375623; font-size:12px; padding:6px;")
        info.setWordWrap(True)

        open_btn = QPushButton("📂  결과 파일 열기")
        open_btn.setStyleSheet(f"""
            QPushButton {{ background:{SUCCESS}; color:white; border-radius:8px;
                padding:10px 24px; font-size:13px; font-weight:bold; }}
            QPushButton:hover {{ background:#4CAF50; }}
        """)
        open_btn.clicked.connect(lambda: os.startfile(out_path) if sys.platform=="win32"
                                  else os.system(f'open "{out_path}"' if sys.platform=="darwin"
                                                 else f'xdg-open "{out_path}"'))

        self.result_layout.addWidget(info)
        self.result_layout.addWidget(open_btn)
        self.result_layout.addStretch()


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor(BG))
    app.setPalette(palette)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
