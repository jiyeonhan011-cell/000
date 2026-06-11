#!/usr/bin/env python3
"""
창고이동검수 자동화 - TCT 케이터링 이동처리(날짜별) 수량 수집
I열(비고)에서 'TCT 케이터링 이동처리' 항목만 필터링하고
F열(이동수량)을 날짜별로 집계하여 결과 엑셀 파일로 저장합니다.
"""

import sys
import os
import re
from pathlib import Path

try:
    import xlrd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("필요한 라이브러리를 설치합니다...")
    os.system("pip install xlrd openpyxl -q")
    import xlrd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


def load_xls(filepath: str):
    """xls 파일을 읽어 데이터 반환"""
    try:
        wb = xlrd.open_workbook(filepath)
    except Exception as e:
        print(f"[오류] 파일을 열 수 없습니다: {e}")
        sys.exit(1)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, j)) for j in range(ws.ncols)]
    rows = []
    for i in range(1, ws.nrows):
        row = {headers[j]: ws.cell_value(i, j) for j in range(ws.ncols)}
        rows.append(row)
    return headers, rows


def extract_date_from_note(note: str):
    """비고에서 날짜 추출 (예: TCT 케이터링 이동처리(2026-06-05) -> 2026-06-05)"""
    m = re.search(r'\((\d{4}-\d{2}-\d{2})\)', note)
    return m.group(1) if m else None


def collect_tct_catering(rows: list):
    """
    I열(비고)에서 'TCT 케이터링 이동처리' 항목만 필터링
    F열(이동수량) 수집
    """
    TARGET = "TCT 케이터링 이동처리"
    filtered = []
    for row in rows:
        note = str(row.get("비고", ""))
        if TARGET in note:
            date = extract_date_from_note(note)
            qty = row.get("이동수량", 0)
            try:
                qty = float(qty)
            except (ValueError, TypeError):
                qty = 0.0
            filtered.append({
                "이동일자": str(row.get("이동일자", "")).strip(),
                "비고날짜": date or str(row.get("이동일자", "")).strip(),
                "이동번호": str(row.get("이동번호", "")).strip(),
                "품목코드": str(row.get("품목코드", "")).strip(),
                "품목명": str(row.get("품목명", "")).strip(),
                "규격": str(row.get("규격", "")).strip(),
                "단위": str(row.get("단위", "")).strip(),
                "이동수량": qty,
                "비고": note.strip(),
            })
    return filtered


def summarize_by_date(filtered: list):
    """날짜별 이동수량 합계"""
    summary = {}
    for item in filtered:
        key = item["비고날짜"]
        summary[key] = summary.get(key, 0) + item["이동수량"]
    return dict(sorted(summary.items()))


def save_result(filtered: list, summary: dict, output_path: str):
    """결과를 xlsx 파일로 저장"""
    wb = openpyxl.Workbook()

    # ── 시트1: 날짜별 합계 ──────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "날짜별합계"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    sum_headers = ["날짜", "이동수량 합계(EA)"]
    for c, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    total_qty = 0
    for r, (date, qty) in enumerate(summary.items(), 2):
        ws_sum.cell(row=r, column=1, value=date).border = border
        ws_sum.cell(row=r, column=1).alignment = center
        ws_sum.cell(row=r, column=2, value=qty).border = border
        ws_sum.cell(row=r, column=2).alignment = center
        total_qty += qty

    # 합계 행
    total_row = len(summary) + 2
    t1 = ws_sum.cell(row=total_row, column=1, value="합 계")
    t2 = ws_sum.cell(row=total_row, column=2, value=total_qty)
    for cell in (t1, t2):
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D6E4F0")
        cell.alignment = center
        cell.border = border

    ws_sum.column_dimensions["A"].width = 15
    ws_sum.column_dimensions["B"].width = 20

    # ── 시트2: 상세 내역 ──────────────────────────────
    ws_det = wb.create_sheet("상세내역")
    det_headers = ["날짜(비고)", "이동일자", "이동번호", "품목코드", "품목명", "규격", "단위", "이동수량", "비고"]
    det_keys   = ["비고날짜",  "이동일자", "이동번호", "품목코드", "품목명", "규격", "단위", "이동수량", "비고"]
    col_widths  = [14, 14, 22, 14, 40, 18, 7, 12, 50]

    for c, h in enumerate(det_headers, 1):
        cell = ws_det.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r, item in enumerate(filtered, 2):
        for c, key in enumerate(det_keys, 1):
            cell = ws_det.cell(row=r, column=c, value=item[key])
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for c, w in enumerate(col_widths, 1):
        ws_det.column_dimensions[get_column_letter(c)].width = w

    wb.save(output_path)
    print(f"저장 완료: {output_path}")


def main():
    if len(sys.argv) < 2:
        # 기본 경로 (같은 폴더 내 xls 자동 탐색)
        search_dir = Path(__file__).parent
        xls_files = list(search_dir.glob("*.xls")) + list(search_dir.glob("*.xlsx"))
        if not xls_files:
            print("사용법: python3 warehouse_tct_catering.py <입력파일.xls>")
            sys.exit(1)
        input_path = str(xls_files[0])
        print(f"파일 자동감지: {input_path}")
    else:
        input_path = sys.argv[1]

    output_path = str(Path(input_path).with_name(
        Path(input_path).stem + "_TCT케이터링이동처리결과.xlsx"
    ))
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]

    print(f"입력: {input_path}")
    headers, rows = load_xls(input_path)
    print(f"전체 데이터: {len(rows)}행")

    filtered = collect_tct_catering(rows)
    print(f"TCT 케이터링 이동처리 건수: {len(filtered)}건")

    if not filtered:
        print("해당 데이터가 없습니다. 비고(I열) 내용을 확인하세요.")
        sys.exit(0)

    summary = summarize_by_date(filtered)
    print("\n[날짜별 이동수량 합계]")
    total = 0
    for date, qty in summary.items():
        print(f"  {date}: {qty:>8.0f} EA")
        total += qty
    print(f"  {'합 계':>12}: {total:>8.0f} EA")

    save_result(filtered, summary, output_path)


if __name__ == "__main__":
    main()
