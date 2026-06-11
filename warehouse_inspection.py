#!/usr/bin/env python3
"""
창고이동검수 자동화
- 창고이동 파일(xls): I열 'TCT 케이터링 이동처리' 항목 / F열 이동수량
- 케이터링 파일(xlsx): K열 수량 (창고이동 2번 → ÷2 처리)
- 날짜+품목명 기준 매칭 → 일치/불일치 리스트업
"""

import sys
import os
import re
from pathlib import Path
from collections import defaultdict

try:
    import xlrd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install xlrd openpyxl -q")
    import xlrd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── 스타일 헬퍼 ─────────────────────────────────────────────
def _cell(ws, row, col, value=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    return c

BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def hdr_style(cell, bg="1F4E79"):
    cell.font = Font(color="FFFFFF", bold=True, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = CENTER
    cell.border = BORDER

def data_style(cell, align=LEFT, bold=False, bg=None):
    cell.font = Font(bold=bold, size=10)
    cell.alignment = align
    cell.border = BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


# ── 이름 정규화 ─────────────────────────────────────────────
def normalize(name: str) -> str:
    # (면세), (대체), (수입) 등 괄호 접두사 제거 후 공백 정리
    name = re.sub(r'^\([^)]+\)\s*', '', name)
    return re.sub(r'\s+', ' ', name).strip()


# ── 창고이동 파일 로드 (TCT 케이터링 이동처리만) ─────────────
def load_warehouse(path: str):
    """날짜+품목명 → 이동수량 합계 dict"""
    try:
        wb = xlrd.open_workbook(path)
    except Exception as e:
        print(f"[오류] 창고이동 파일 오류: {e}"); sys.exit(1)
    ws = wb.sheet_by_index(0)
    TARGET = "TCT 케이터링 이동처리"
    result = defaultdict(float)   # (date, norm_name) -> qty
    details = {}                  # (date, norm_name) -> {name, code, spec, unit}

    for i in range(1, ws.nrows):
        note = str(ws.cell_value(i, 8))
        if TARGET not in note:
            continue
        m = re.search(r'\((\d{4}-\d{2}-\d{2})\)', note)
        date = m.group(1) if m else str(ws.cell_value(i, 3)).strip()
        name = str(ws.cell_value(i, 10)).strip()
        qty  = float(ws.cell_value(i, 5) or 0)
        key  = (date, normalize(name))
        result[key] += qty
        if key not in details:
            details[key] = {
                "품목코드": str(ws.cell_value(i, 9)).strip(),
                "품목명":   name,
                "규격":     str(ws.cell_value(i, 11)).strip(),
                "단위":     str(ws.cell_value(i, 12)).strip(),
            }
    return result, details


# ── 케이터링 파일 로드 ────────────────────────────────────────
def load_catering(path: str):
    """날짜+제품명 → 수량합계÷2 dict"""
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"[오류] 케이터링 파일 오류: {e}"); sys.exit(1)
    ws = wb.active
    result  = defaultdict(float)
    details = {}

    for i in range(4, ws.max_row + 1):
        date = ws.cell(i, 3).value
        if not date or str(date) == "센터명":
            continue
        name = str(ws.cell(i, 8).value or "").strip()
        qty  = ws.cell(i, 11).value or 0
        if not name:
            continue
        key = (str(date), normalize(name))
        result[key] += float(qty)
        if key not in details:
            details[key] = {
                "제품코드": str(ws.cell(i, 7).value or "").strip(),
                "제품명":   name,
                "규격":     str(ws.cell(i, 9).value or "").strip(),
                "단위":     str(ws.cell(i, 10).value or "").strip(),
            }

    # 창고이동 2회 → ÷2
    halved = {k: v / 2 for k, v in result.items()}
    return halved, details


# ── 검수 비교 ─────────────────────────────────────────────────
def inspect(wh_qty, wh_det, cat_qty, cat_det):
    """
    모든 키를 순회하여 3가지 유형으로 분류:
      - 일치: 양쪽 있고 수량 동일
      - 수량불일치: 양쪽 있으나 수량 다름
      - 창고이동만: 창고이동 파일에만 존재
      - 케이터링만: 케이터링 파일에만 존재
    """
    all_keys = set(wh_qty.keys()) | set(cat_qty.keys())
    matched       = []
    qty_mismatch  = []   # 양쪽 있으나 수량 다름
    wh_only       = []   # 창고이동에만 있음
    cat_only      = []   # 케이터링에만 있음

    for key in sorted(all_keys):
        date, norm_name = key
        wq = wh_qty.get(key, None)
        cq = cat_qty.get(key, None)

        wdet = wh_det.get(key, {})
        cdet = cat_det.get(key, {})

        item_name = wdet.get("품목명") or cdet.get("제품명") or norm_name
        wcode = wdet.get("품목코드", "-")
        ccode = cdet.get("제품코드", "-")
        spec  = wdet.get("규격") or cdet.get("규격", "-")
        unit  = wdet.get("단위") or cdet.get("단위", "-")

        row = {
            "날짜":            date,
            "품목명":          item_name,
            "품목코드":        wcode,
            "제품코드":        ccode,
            "규격":            spec,
            "단위":            unit,
            "창고이동수량":    wq,
            "케이터링수량(÷2)": cq,
            "차이":            (cq or 0) - (wq or 0),
        }
        if wq is not None and cq is not None:
            if abs(wq - cq) < 0.001:
                matched.append(row)
            else:
                qty_mismatch.append(row)
        elif wq is not None:
            wh_only.append(row)
        else:
            cat_only.append(row)

    return matched, qty_mismatch, wh_only, cat_only


# ── xlsx 저장 ─────────────────────────────────────────────────
def save_result(matched, qty_mismatch, wh_only, cat_only, output_path: str):
    wb = openpyxl.Workbook()
    total = len(matched) + len(qty_mismatch) + len(wh_only) + len(cat_only)

    # ── 요약 시트 ─────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "검수요약"

    summary_data = [
        ("전체 비교 품목 수 (날짜+품목명)",         total),
        ("✅ 일치 항목",                            len(matched)),
        ("❌ 수량 불일치 (양쪽 존재, 수량 다름)",    len(qty_mismatch)),
        ("⚠️ 창고이동에만 있는 항목",               len(wh_only)),
        ("⚠️ 케이터링에만 있는 항목",               len(cat_only)),
        ("일치율 (공통항목 기준)",
         f"{len(matched)/(len(matched)+len(qty_mismatch))*100:.1f}%"
         if (matched or qty_mismatch) else "-"),
    ]
    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 18

    ws_sum.merge_cells("A1:B1")
    t = ws_sum.cell(row=1, column=1, value="창고이동검수 요약")
    hdr_style(t, "1F4E79")

    for r, (label, val) in enumerate(summary_data, 2):
        c1 = ws_sum.cell(row=r, column=1, value=label)
        c2 = ws_sum.cell(row=r, column=2, value=val)
        data_style(c1, LEFT)
        data_style(c2, CENTER)

    # ── 수량불일치 시트 ───────────────────────────────────────
    _write_detail_sheet(wb, "수량불일치", qty_mismatch, bg_header="C00000", row_bg="FCE4D6")

    # ── 창고이동에만 있는 시트 ────────────────────────────────
    _write_detail_sheet(wb, "창고이동만_미매칭", wh_only, bg_header="843C0C", row_bg="FFF2CC")

    # ── 케이터링에만 있는 시트 ────────────────────────────────
    _write_detail_sheet(wb, "케이터링만_미매칭", cat_only, bg_header="7030A0", row_bg="EAD1F5")

    # ── 일치 시트 ─────────────────────────────────────────────
    _write_detail_sheet(wb, "일치항목", matched, bg_header="375623", row_bg="E2EFDA")

    wb.save(output_path)
    print(f"저장 완료: {output_path}")


def _write_detail_sheet(wb, title, rows, bg_header="1F4E79", row_bg=None, extra_note=None):
    ws = wb.create_sheet(title)
    headers = ["날짜", "품목명", "품목코드(창고)", "제품코드(케이터링)", "규격", "단위",
               "창고이동수량(F열)", "케이터링수량(K열÷2)", "차이"]
    keys    = ["날짜", "품목명", "품목코드", "제품코드", "규격", "단위",
               "창고이동수량", "케이터링수량(÷2)", "차이"]
    widths  = [13, 38, 16, 18, 18, 7, 18, 20, 10]

    for c, h in enumerate(headers, 1):
        cell = _cell(ws, 1, c, h)
        hdr_style(cell, bg_header)

    for r, item in enumerate(rows, 2):
        for c, key in enumerate(keys, 1):
            val = item.get(key)
            if val is None:
                val = "-"
            cell = _cell(ws, r, c, val)
            align = CENTER if c in (1, 6, 7, 8, 9) else LEFT
            bg = row_bg if (r % 2 == 0) else None
            # 차이 셀 색상
            if key == "차이" and isinstance(val, float) and abs(val) > 0.001:
                bg = "FFE0E0"
            data_style(cell, align, bg=bg)
            if key == "차이" and isinstance(val, float):
                cell.number_format = '+0.##;-0.##;0'

    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── 메인 ──────────────────────────────────────────────────────
def main():
    default_wh  = "/root/.claude/uploads/4b31195f-0362-5c7f-baf9-3f04e5c8214d/21498ca7-20260608133431.xls"
    default_cat = "/root/.claude/uploads/4b31195f-0362-5c7f-baf9-3f04e5c8214d/544784f8-_____20260608_1420.xlsx"

    wh_path  = sys.argv[1] if len(sys.argv) > 1 else default_wh
    cat_path = sys.argv[2] if len(sys.argv) > 2 else default_cat
    out_path = sys.argv[3] if len(sys.argv) > 3 else "창고이동검수결과.xlsx"

    print(f"창고이동 파일: {wh_path}")
    print(f"케이터링 파일:  {cat_path}")

    wh_qty,  wh_det  = load_warehouse(wh_path)
    cat_qty, cat_det = load_catering(cat_path)

    print(f"창고이동 항목 수 (날짜+품목): {len(wh_qty)}")
    print(f"케이터링 항목 수 (날짜+품목, ÷2 적용): {len(cat_qty)}")

    matched, qty_mismatch, wh_only, cat_only = inspect(wh_qty, wh_det, cat_qty, cat_det)

    total = len(matched) + len(qty_mismatch) + len(wh_only) + len(cat_only)
    common = len(matched) + len(qty_mismatch)
    print(f"\n[검수 결과]")
    print(f"  전체: {total}건")
    print(f"  ✅ 일치: {len(matched)}건  ({len(matched)/common*100:.1f}%, 공통항목 기준)" if common else "")
    print(f"  ❌ 수량불일치: {len(qty_mismatch)}건  ({len(qty_mismatch)/common*100:.1f}%)" if common else "")
    print(f"  ⚠️  창고이동에만: {len(wh_only)}건")
    print(f"  ⚠️  케이터링에만: {len(cat_only)}건")

    if qty_mismatch:
        print("\n[수량불일치 샘플 (최대 10건)]")
        for item in qty_mismatch[:10]:
            wq = item['창고이동수량']
            cq = item['케이터링수량(÷2)']
            print(f"  {item['날짜']} | {item['품목명'][:30]:<30} | 창고이동={wq} | 케이터링÷2={cq} | 차이={item['차이']}")

    save_result(matched, qty_mismatch, wh_only, cat_only, out_path)


if __name__ == "__main__":
    main()
